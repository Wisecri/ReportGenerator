####################################################################################################
# Utils.py
import pandas as pd
import re
import warnings
import os
import numpy as np
import matplotlib.pyplot as plt
from fpdf import FPDF
from optparse import OptionParser
from PIL import Image, ImageTk
import io
import PySimpleGUI as sg
import base64
from matplotlib.patches import Patch
import openpyxl

# Define your custom theme
custom_theme = {
    "BACKGROUND": "#83142c",  # Background color
    "TEXT": "#FFFFFF",  # Text color
    "INPUT": "#F0F0F0",  # Input field color
    "TEXT_INPUT": "#000000",  # Text color within input fields
    "SCROLL": "#F0F0F0",  # Scrollbar color
    "BUTTON": ("#000000", "#F0F0F0"),  # Button text and background colors
    "PROGRESS": ("#000000", "#F0F0F0"),  # Progress bar colors
    "BORDER": 0,  # Border width
    "SLIDER_DEPTH": 0,  # Slider depth
    "PROGRESS_DEPTH": 0,  # Progress bar depth
}

# Set the custom theme
sg.theme_add_new("CustomTheme", custom_theme)
sg.theme("CustomTheme")

def get_img_data(f, maxsize=(256, 255), first=False):
    img = Image.open(f)
    img.thumbnail(maxsize)
    if first:                     # Tkinter is inactive the first time
        bio = io.BytesIO()
        img.save(bio, format='PNG')
        del img
        return bio.getvalue()
    return ImageTk.PhotoImage(img)

def clean(text, cap=True, override=False):
    parts = re.split(r"(\([^)]*\))", text)
    for i in range(len(parts)):
        if not parts[i].startswith("(") and not parts[i].endswith(")"):
            if cap:
                parts[i] = parts[i].upper()
        elif not override:
            parts[i] = ""
    return "".join(parts)

####################################################################################################
# Dictionaries

# Corsi di laurea
cdls_dict = {
    "Acquacoltura e igiene delle produzioni ittiche (Cesenatico)": '1',
    "Antropologia, religioni, civiltà orientali (Bologna)": '2',
    "Architettura (Cesena)": '3',
    "Architettura - Ingegneria (Bologna)": '4',
    "Astronomia (Bologna)": '5',
    "Beni culturali (Bologna)" : '6',
    "Biotecnologie (Bologna)" : '7',
    "Building construction engineering (Bologna)": '8',
    "Business and economics (Bologna)": '9',
    "Chimica e chimica dei materiali (Bologna)": '10',
    "Chimica e tecnologie farmaceutiche (Bologna)": '11',
    "Chimica e tecnologie per l'ambiente e per i materiali (Faenza)": '12',
    "Chimica e tecnologie per l'ambiente e per i materiali (Rimini)": '13',
    "Chimica industriale (Bologna)": '14',
    "Conservazione e restauro dei beni culturali (Bologna)": '15',
    "Consulente del lavoro e delle relazioni aziendali (Bologna)": '16',
    "Culture e pratiche della moda (Rimini)": '17',
    "DAMS - discipline delle arti, della musica e dello spettacolo (Bologna)": '18',
    "Design del prodotto industriale (Bologna)": '19',
    "Dietistica (Bologna)": '20',
    "Economia aziendale (Bologna)": '21',
    "Economia del turismo (Rimini) - curriculum economia del turismo (ita)- curriculum international torism and leisure industries (eng)": '22',
    "Economia dell'impresa (Rimini) -curriculum percorso manageriale/professionale (ita) - curriculum financial and business management (eng)" : '23',
    "Economia e commercio - curriculum amministrativo finanza e controllo (Bologna)": '24',
    "Economia e marketing nel sistema agro-industriale (Bologna)": '25',
    "Economia e mercati e istituzioni (Bologna)": '26',
    "Economics and finance (Bologna)": '27',
    "Economics, politics and social sciences (Bologna)": '28',
    "Educatore nei servizi per l'infanzia (Bologna)": '29',
    "Educatore sociale e culturale (Bologna)": '30',
    "Educatore sociale e culturale (Rimini)": '31',
    "Educazione professionale (Bologna)": '32',
    "European studies (Forlì)": '33',
    "Farmacia (Bologna)": '34',
    "Filosofia (Bologna)": '35',
    "Fisica (Bologna)": '36',
    "Fisioterapia (Bologna)": '37',
    "Genomics (Bologna)" : '38',
    "Giurisprudenza (Bologna)": '39',
    "Giurisprudenza (Ravenna)": '40',
    "Giurista per le imprese e per la pubblica amministrazione (Bologna)": '41',
    "Igiene dentale (Bologna)": '42',
    "Infiermieristica (Bologna)": '43',
    "Infermieristica (Faenza)": '44',
    "Infermieristica (Rimini)": '45',
    "Informatica (Bologna)": '46',
    "Informatica per il management (Bologna)": '47',
    "Ingegneria aerospaziale (Forlì)": '48',
    "Ingegneria biomedica (Bologna)": '49',
    "Ingegneria chimica e biochimica (Bologna)": '50',
    "Ingegneria civile (Bologna)": '51',
    "Ingegneria dell'automazione (Bologna)": '52',
    "Ingegneria e scienze informatiche (Cesena)": '53',
    "Ingegneria elettronica (Bologna)": '54',
    "Ingegneria elettronica e telecomunicazioni (Bologna)": '55',
    "Ingegneria dell'energia elettrica (Bologna)": '56',
    "Ingegneria energetica (Bologna)": '57',
    "Ingegneria gestionale (Bologna)": '58',
    "Ingegneria informatica (Bologna)": '59',
    "Ingegneria meccanica (Bologna)": '60',
    "Ingegneria meccanica (Forlì)": '61',
    "Ingegneria per l'ambiente e il territorio (Bologna)": '62',
    "International studies (Forlì)": '63',
    "Lettere (Bologna)": '64',
    "Lingue e letterature straniere (Bologna)": '65',
    "Lingue e tecnologie per la comunicazione interculturale (Bologna)": '66',
    "Lingue, mercati e culture dell'Asia e dell'Africa mediterranea (Bologna)": '67',
    "Logopedia (Bologna)": '68',
    "Management and economics (Bologna)": '69',
    "Management e marketing (Bologna)": '70',
    "Matematica (Bologna)": '71',
    "Meccatronica (Bologna)": '72',
    "Medicina e chirurgia (Bologna)": '73',
    "Medicina e chirurgia (Forlì)": '74',
    "Medicina e chirurgia (Ravenna)": '75',
    "Medicina veterinaria (Bologna)": '76',
    "Medicine and surgery (Bologna)": '77',
    "Metodologie chimiche per prodotti e processi (Bologna)": '78',
    "Odontoiatria e protesi dentaria (Bologna)": '79',
    "Ostetricia (Bologna)": '80',
    "Pharmacy (Rimini)": '81',
    "Podologia (Bologna)": '82',
    "Produzioni animali (Bologna)": '83',
    "Scienza dei materiali (Bologna)" : '84',
    "Scienze ambientali (Bologna)": '85',
    "Scienze biologiche (Bologna)": '86',
    "Scienze della comunicazione (Bologna)": '87',
    "Scienze della formazione primaria (Bologna)": '88',
    "Scienze delle attività motorie e sportive (Bologna)": '89',
    "Scienze delle attività motorie e sportive (Rimini)": '90',
    "Scienze e culture della gastronomia (Cesena)": '91',
    "Scienze e tecniche psicologiche (Bologna)": '92',
    "Scienze e tecnologie per il verde e il paesaggio (Bologna)": '93',
    "Scienze farmaceutiche applicate (Bologna)": '94',
    "Scienze geologiche (Bologna)": '95',
    "Scienze internazionali e diplomatiche (Forlì) - curriculum scienze internazionali e diplomatiche (ita) - curriculum diplomatic and international sciences (eng)" : '96',
    "Scienze naturali (Bologna)": '97',
    "Scienze politiche, sociali e internazionali (Bologna)": '98',
    "Scienze statistiche (Bologna)": '99',
    # "Scienze statistiche (curricula ita: economia, impresa, biodemografico) (Bologna)" : '99',
    # "Scienze statistiche (curricula stats and maths) (Bologna)" : '100',
    "Servizio sociale (Bologna)": '101',
    "Sociologia (Forlì)": '102',
    "Statistica, finanza e assicurazioni (Rimini)": '103',
    "Storia (Bologna)": '104',
    "Storia, società e culture del Mediterraneo (Bologna)": '105',
    "Sviluppo e cooperazione internazionale (Bologna)": '106',
    "Tecniche della prevensione nell'ambiente e nei luoghi di lavoro (Bologna)": '107',
    "Tecniche di laboratorio biomedico (Bologna)": '108',
    "Tecniche di neurofisiopatologia (Bologna)": '109',
    "Tecniche di radiologia medica, per immagini e radioterapia (Bologna)": '110',
    "Tecniche ortopediche (Bologna)": '111',
    "Tecniche per l'edilizia e il territorio (Bologna)": '112',
    "Tecnologie agrarie (Bologna)": '113',
    "Tecnologie alimentari (Cesena)": '114',
    "Tecnologie dei sistemi informatici (Bologna)": '115',
    "Tecnologie per il territorio e l'ambiente agro-forestale (Bologna)": '116',
    "Viticoltura ed enologia (Cesena)": '117'
}

# Dipartimenti
dips_dict = {
    "DIMEVET" : ["Scienze mediche veterinarie", "1"],
    "DiSCi" : ["Storia, Culture e Civiltà", "2"],
    "DA" : ["Architettura", "3"],
    "DIFA" : ["Fisica e Astronomia 'Augusto Righi'", "4"],
    "DBC" : ["Beni Culturali", "5"],
    "FaBiT" : ["Farmacia e Biotecnologie", "6"],
    "DICAM" : ["Ingegneria civile, chimica, ambientale e dei materiali", "7"],
    "DiSA" : ["Scienze aziendali", "8"],
    "CHIM" : ["Chimica 'Giacomo Ciamician'", "9"],
    "CHIMIND" : ["Chimica Industriale 'Toso Montanari'", "10"],
    "DSG" : ["Scienze Giuridiche", "11"],
    "DAR" : ["Delle Arti", "12"],
    "DIMEC" : ["Scienze Mediche e Chirurgiche", "13"],
    "DSE" : ["Scienze Economiche", "14"],
    "DISTAL" : ["Scienze e Tecnologie Agro-Alimentari", "15"],
    "EDU" : ["Scienze dell'Educazione 'Giovanni Maria Bertin'", "16"],
    "DIBINEM" : ["Scienze biomediche e neuromotorie", "17"],
    "SPS" : ["Scienze Politiche e Sociali", "18"],
    "FILCOM" : ["Filosofia e Comunicazione", "19"],
    "DISI" : ["Informatica, Scienza e Ingegneria", "20"],
    "DIN" : ["Ingegneria industriale", "21"],
    "DEI" : ["Ingegneria dell'Energia Elettrica e dell'Informazione 'Guglielmo Marconi'", "22"],
    "FICLIT" : ["Filologia Classica e Italianistica", "23"],
    "LILEC" : ["Lingue, Letterature e Culture Moderne", "24"],
    "DIT" : ["Interpretazione e Traduzione", "25"],
    "MAT" : ["Matematica", "26"],
    "BiGeA" : ["Scienze Biologiche, Geologiche ed Ambientali", "27"],
    "QUVI" : ["Scienze per la Qualità della Vita", "28"],
    "PSI" : ["Psicologia 'Renzo Canestrari'", "29"],
    "STAT" : ["Scienze Statistiche 'Paolo Fortunati'", "30"],
    "SDE" : ["Scienze del Diritto e dell'Economia", "31"]
}

# Corsi di laurea - Dipartimenti
cdl_dips_amb_dict = {
    "Acquacoltura e igiene delle produzioni ittiche (Cesenatico)": ["DIMEVET", "Medicina Veterinaria"],
    "Antropologia, religioni, civiltà orientali (Bologna)": ["DiSCi", "Studi Umanistici"],
    "Architettura (Cesena)": ["DA", "Ingegneria e Architettura"],
    "Architettura - Ingegneria (Bologna)": ["DA", "Ingegneria e Architettura"],
    "Astronomia (Bologna)": ["DIFA", "Scienze"],
    "Beni culturali (Bologna)": ["DBC", "Studi Umanistici"],
    "Biotecnologie (Bologna)": ["FaBiT", "Farmacie e biotecnologie"],
    "Building construction engineering (Bologna)": ["DICAM", "Ingegneria e Architettura"],
    "Business and economics (Bologna)": ["DiSA", "Economia e Management"],
    "Chimica e chimica dei materiali (Bologna)": ["CHIM", "Scienze"],
    "Chimica e tecnologie farmaceutiche (Bologna)": ["FaBiT", "Farmacie e biotecnologie"],
    "Chimica e tecnologie per l'ambiente e per i materiali (Faenza)": ["CHIMIND", "Scienze"],
    "Chimica e tecnologie per l'ambiente e per i materiali (Rimini)": ["CHIMIND", "Scienze"],
    "Chimica industriale (Bologna)": ["CHIMIND", "Scienze"],
    "Conservazione e restauro dei beni culturali (Bologna)": ["DBC", "Studi Umanistici"],
    "Consulente del lavoro e delle relazioni aziendali (Bologna)": ["DSG", "Giurisprudenza"],
    "Culture e pratiche della moda (Rimini)": ["DAR", "Studi Umanistici"],
    "DAMS - discipline delle arti, della musica e dello spettacolo (Bologna)": ["DAR", "Studi Umanistici"],
    "Design del prodotto industriale (Bologna)": ["DA", "Ingegneria e Architettura"],
    "Dietistica (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Economia aziendale (Bologna)": ["DiSA", "Economia e Management"],
    "Economia del turismo (Rimini) - curriculum economia del turismo (ita)- curriculum international torism and leisure industries (eng)": ["DSE", "Economia e Management"],
    "Economia dell'impresa (Rimini) -curriculum percorso manageriale/professionale (ita) - curriculum financial and business management (eng)": ["DiSA", "Economia e Management"],
    "Economia e commercio - curriculum amministrativo finanza e controllo (Bologna)": ["DiSA", "Economia e Management"],
    "Economia e marketing nel sistema agro-industriale (Bologna)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Economia e mercati e istituzioni (Bologna)": ["DSE", "Economia e Management"],
    "Economics and finance (Bologna)": ["DSE", "Economia e Management"],
    "Economics, politics and social sciences (Bologna)": ["DSE", "Economia e Management"],
    "Educatore nei servizi per l'infanzia (Bologna)": ["EDU", "Scienze dell'Educazione e della Formazione"],
    "Educatore sociale e culturale (Bologna)": ["EDU", "Scienze dell'Educazione e della Formazione"],
    "Educatore sociale e culturale (Rimini)": ["EDU", "Scienze dell'Educazione e della Formazione"],
    "Educazione professionale (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "European studies (Forlì)": ["SPS", "Scienze Poltiche"],
    "Farmacia (Bologna)": ["FaBiT", "Farmacie e biotecnologie"],
    "Filosofia (Bologna)": ["FILCOM", "Studi Umanistici"],
    "Fisica (Bologna)": ["DIFA", "Scienze"],
    "Fisioterapia (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Genomics (Bologna)": ["FaBiT", "Farmacie e biotecnologie"],
    "Giurisprudenza (Bologna)": ["DSG", "Giurisprudenza"],
    "Giurisprudenza (Ravenna)": ["DSG", "Giurisprudenza"],
    "Giurista per le imprese e per la pubblica amministrazione (Bologna)": ["DSG", "Giurisprudenza"],
    "Igiene dentale (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Infiermieristica (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Infermieristica (Faenza)": ["DIMEC", "Medicina e Chirurgia"],
    "Infermieristica (Rimini)": ["DIMEC", "Medicina e Chirurgia"],
    "Informatica (Bologna)": ["DISI", "Scienze"],
    "Informatica per il management (Bologna)": ["DISI", "Scienze"],
    "Ingegneria aerospaziale (Forlì)": ["DIN", "Ingegneria e Architettura"],
    "Ingegneria biomedica (Bologna)": ["DEI", "Ingegneria e Architettura"],
    "Ingegneria chimica e biochimica (Bologna)": ["DICAM", "Ingegneria e Architettura"],
    "Ingegneria civile (Bologna)": ["DICAM", "Ingegneria e Architettura"],
    "Ingegneria dell'automazione (Bologna)": ["DEI", "Ingegneria e Architettura"],
    "Ingegneria e scienze informatiche (Cesena)": ["DISI", "Ingegneria e Architettura"],
    "Ingegneria elettronica (Bologna)": ["DEI", "Ingegneria e Architettura"],
    "Ingegneria elettronica e telecomunicazioni (Bologna)": ["DEI", "Ingegneria e Architettura"],
    "Ingegneria dell'energia elettrica (Bologna)": ["DEI", "Ingegneria e Architettura"],
    "Ingegneria energetica (Bologna)": ["DIN", "Ingegneria e Architettura"],
    "Ingegneria gestionale (Bologna)": ["DIN", "Ingegneria e Architettura"],
    "Ingegneria informatica (Bologna)": ["DISI", "Ingegneria e Architettura"],
    "Ingegneria meccanica (Bologna)": ["DIN", "Ingegneria e Architettura"],
    "Ingegneria meccanica (Forlì)": ["DIN", "Ingegneria e Architettura"],
    "Ingegneria per l'ambiente e il territorio (Bologna)": ["DICAM", "Ingegneria e Architettura"],
    "International studies (Forlì)": ["SPS", "Scienze Poltiche"],
    "Lettere (Bologna)": ["FICLIT", "Studi Umanistici"],
    "Lingue e letterature straniere (Bologna)": ["LILEC", "Lingue e letterature, traduzione e interpretazione"],
    "Lingue e tecnologie per la comunicazione interculturale (Bologna)": ["LILEC", "Lingue e letterature, traduzione e interpretazione"],
    "Lingue, mercati e culture dell'Asia e dell'Africa mediterranea (Bologna)": ["LILEC", "Lingue e letterature, traduzione e interpretazione"],
    "Logopedia (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Management and economics (Bologna)": ["DiSA", "Economia e Management"],
    "Management e marketing (Bologna)": ["DiSA", "Economia e Management"],
    "Matematica (Bologna)": ["MAT", "Scienze"],
    "Meccatronica (Bologna)": ["DEI", "Ingegneria e Architettura"],
    "Medicina e chirurgia (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Medicina e chirurgia (Forlì)": ["DIMEC", "Medicina e Chirurgia"],
    "Medicina e chirurgia (Ravenna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Medicina veterinaria (Bologna)": ["DIMEVET", "Medicina Veterinaria"],
    "Medicine and surgery (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Metodologie chimiche per prodotti e processi (Bologna)": ["CHIMIND", "Scienze"],
    "Odontoiatria e protesi dentaria (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Ostetricia (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Pharmacy (Rimini)": ["FaBiT", "Farmacie e biotecnologie"],
    "Podologia (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Produzioni animali (Bologna)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Scienza dei materiali (Bologna)": ["BiGeA", "Scienze"],
    "Scienze ambientali (Bologna)": ["BiGeA", "Scienze"],
    "Scienze biologiche (Bologna)": ["BiGeA", "Scienze"],
    "Scienze della comunicazione (Bologna)": ["DAR", "Studi Umanistici"],
    "Scienze della formazione primaria (Bologna)": ["EDU", "Scienze dell'Educazione e della Formazione"],
    "Scienze delle attività motorie e sportive (Bologna)": ["QUVI", "Scienze motorie"],
    "Scienze delle attività motorie e sportive (Rimini)": ["QUVI", "Scienze motorie"],
    "Scienze e culture della gastronomia (Cesena)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Scienze e tecniche psicologiche (Bologna)": ["PSI", "Psiocologia"],
    "Scienze e tecnologie per il verde e il paesaggio (Bologna)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Scienze farmaceutiche applicate (Bologna)": ["FaBiT", "Farmacie e biotecnologie"],
    "Scienze geologiche (Bologna)": ["BiGeA", "Scienze"],
    "Scienze internazionali e diplomatiche (Forlì) - curriculum scienze internazionali e diplomatiche (ita) - curriculum diplomatic and international sciences (eng)": ["SPS", "Scienze Poltiche"],
    "Scienze naturali (Bologna)": ["BiGeA", "Scienze"],
    "Scienze politiche, sociali e internazionali (Bologna)": ["SPS", "Scienze Poltiche"],
    "Scienze statistiche (Bologna)": ["STAT", "Scienze Statistiche"],
    # "Scienze statistiche (curricula ita: economia, impresa, biodemografico) (Bologna)": ["STAT", "Scienze Statistiche"],
    # "Scienze statistiche (curricula stats and maths) (Bologna)": ["STAT", "Scienze Statistiche"],
    "Servizio sociale (Bologna)": ["SDE", "Sociologia"],
    "Sociologia (Forlì)": ["SDE", "Sociologia"],
    "Statistica, finanza e assicurazioni (Rimini)": ["STAT", "Scienze Statistiche"],
    "Storia (Bologna)": ["DiSCi", "Studi Umanistici"],
    "Storia, società e culture del Mediterraneo (Bologna)": ["DBC", "Studi Umanistici"],
    "Sviluppo e cooperazione internazionale (Bologna)": ["STAT", "Scienze Poltiche"],
    "Tecniche della prevensione nell'ambiente e nei luoghi di lavoro (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Tecniche di laboratorio biomedico (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Tecniche di neurofisiopatologia (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Tecniche di radiologia medica, per immagini e radioterapia (Bologna)": ["DIMEC", "Medicina e Chirurgia"],
    "Tecniche ortopediche (Bologna)": ["DIBINEM", "Medicina e Chirurgia"],
    "Tecniche per l'edilizia e il territorio (Bologna)": ["DICAM", "Ingegneria e Architettura"],
    "Tecnologie agrarie (Bologna)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Tecnologie alimentari (Cesena)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Tecnologie dei sistemi informatici (Bologna)": ["DISI", "Ingegneria e Architettura"],
    "Tecnologie per il territorio e l'ambiente agro-forestale (Bologna)": ["DISTAL", "Scienze Agro-Alimentari"],
    "Viticoltura ed enologia (Cesena)": ["DISTAL", "Scienze Agro-Alimentari"]
}

# Corsi di laurea POT
pot_cdl_dict = {
    "Gestione degli spazi verdi, dei boschi e delle aree protette (Bari)": '2',
    "Scienze Animali (Bari)": '3',
    "Scienze delle Produzioni e delle Risorse del Mare (Bari)": '4',
    "Scienze e Tecnologie Agrarie": '5',
    "Scienze e Tecnologie Alimentari": '6',
    "Scienze per la valorizzazione del patrimonio gastronomico (Bari)": '7',
    "Tecniche per l'agricoltura sostenibile (Bari)": '8',
    "Scienze Agrarie": '10',
    "Scienze Forestali e Ambientali": '11',
    "Tecnologie Alimentari": '12',
    "Acquacoltura e igiene delle produzioni ittiche (Bologna)": '14',
    "Economia e Marketing nel sistema agro-industriale (Bologna)": '15',
    "Produzioni Animali (Bologna)": '16',
    "Scienze e cultura della gastronomia (Bologna)": '17',
    "Scienze e tecnologie per il verde e il paesaggio (Bologna)": '18',
    "Tecnologie Agrarie (Bologna)": '19',
    "Tecnologie Alimentari (Bologna)":'20', # codice 12
    "Tecnologie per il territorio e l'ambiente agro-forestale (Bologna)": '21',
    "Viticoltura ed Enologia": '22',
    "Sistemi agricoli sostenibili (Brescia)": '24',
    "Scienze Gastronomiche": '26',
    "Tecnico del benessere animale e delle produzioni (Camerino)": '27',
    "Scienze Agrarie e Forestali (Campania)": '29',
    "Scienze e Tecnologie Agrarie (Catania)": '31',
    "Scienze e Tecnologie Alimentari (Catania)": '32',
    "Scienze e tecnologie per la ristorazione e distribuzione degli alimenti mediterranei (Catania)": '33',
    "Tecnologie agrarie e acquacoltura del Delta (Ferrara)": '35',
    "Cultura e Sostenibilità dell'Enogastronomia (Foggia)": '37',
    "Scienze e Tecnologie Agrarie (Foggia)": '38', # codice 5
    "Scienze e Tecnologie Alimentari (Foggia)": '39', # codice 6
    "Scienze Agrarie (Firenze)": '41',
    "Scienze e Tecnologie per la Gestione degli Spazi Verdi e del Paesaggio (Firenze)": '42',
    "Scienze Faunistiche (Firenze)": '43',
    "Scienze Forestali e Ambientali (Firenze)": '44', # codice 11
    "Tecnologie Alimentari (Firenze)": '45', # codice 12
    "Tecnologie e trasformazioni avanzate nel settore legno, arredo, edilizia (Firenze)": '46',
    "Viticoltura ed Enologia (Firenze)": '47',
    "Scienze tecnologie e sicurezza delle produzioni animali (Messina)": '49',
    "Scienze e tecnologie agrarie per la transizione ecologica (Messina)": '50',
    "Agricoltura Sostenibile (Milano)": '52',
    "Allevamento e benessere degli animali d'affezione (Milano)": '53',
    "Produzione e protezione delle piante e dei sistemi del verde (Milano)": '54',
    "Scienze della ristorazione e distribuzione degli alimenti (Milano)": '55',
    "Scienze delle produzioni animali (Milano)": '56',
    "Scienze e Tecnologie per alimenti sostenibili (Milano)": '57',
    "Sistemi digitali in agricoltura (Milano)": '58',
    "Valorizzazione e tutela dell'ambiente e del territorio montano (Milano)": '59',
    "Viticoltura ed Enologia (Milano)": '60',
    "Scienze e Tecnologie agrarie e degli alimenti (Modena)": '62',
    "Scienze e Tecnologie Agrarie e Forestali (Molise)": '64',
    "Scienze e Tecnologie Alimentari (Molise)": '65', # codice 6
    "Scienze Agrarie, Forestali e Ambientali (Napoli)": '67',
    "Scienze Gastronomiche Mediterranee (Napoli)": '68',
    "Tecnologie delle Produzioni animali (Vecchio Ordinamento) e Gestione degli Animali e delle Produzioni (Nuovo Ordinamento) (Napoli)": '69',
    "Tecnologie Alimentari (Napoli)": '70', # codice 12
    "Viticoltura ed Enologia (Napoli)": '71', # codice 22
    "Agroingegneria (Palermo)": '73',
    "Scienze e Tecnologie Agrarie (Palermo)": '74', # codice 5
    "Scienze e Tecnologie Agroalimentari": '75',
    "Scienze Forestali e Ambientali (Palermo)": '76', # codice 11
    "Scienze Gastronomiche (Palermo)": '77', # codice 26
    "Sistemi Agricoli Mediterranei (Palermo)": '78',
    "Viticoltura ed Enologia (Palermo)": '79', # codice 22
    "Animal Care - Tutela del benessere animale (Padova)":'81',
    "Produzioni Biologiche Vegetali (Padova)": '82',
    "Scienze e Cultura della Gastronomia (Padova)": '83',
    "Scienze e Tecnologie Agrarie (Padova)": '84', # codice 5
    "Scienze e Tecnologie Alimentari (Padova)": '85', # codice 6
    "Scienze e Tecnologie Animali (Padova)": '86',
    "Scienze e Tecnologie Viticole ed Enologiche (Padova)": '87',
    "Scienze Tecnico Assistenziali Veterinarie (Padova)": '88',
    "Sicurezza Igienico-Sanitaria degli Alimenti (Padova)": '89',
    "Tecnologie Forestali e Ambientali (Padova)": '90',
    "Scienze Agrarie e Ambientali (Perugia)": '92',
    "Economia e cultura dell'alimentazione (Perugia)": '93',
    "Produzioni Animali (Perugia)": '94', # codice 16
    "Scienze e Tecnologie Agroalimentari (Perugia)": '95', # codice 75
    "Scienze Agrarie (Pisa)": '97', # codice 10
    "Viticoltura ed Enologia (Pisa)": '98', # codice 22 # se nell'elenco si cambia la e
    "Scienze e Tecnologie delle Produzioni Animali (Pisa)": '99',
    "Scienze e Tecnologie Agrarie (Marche)": '101', # codice 5
    "Scienze e Tecnologie Alimentari (Marche)": '102', # codice 6
    "Scienze Forestali e Ambientali (Marche)": '103',
    "Sistemi Agricoli Innovativi (Marche)": '104',
    "Scienze e Tecnologie Alimentari (Parma)": '106', # codice 6
    "Scienze Gastronomiche (Parma)": '107', # codice 26
    "Scienze Zootecniche e Tecnologie delle Produzioni Animali (Parma)": '108',
    "Tecnologie e gestione dell’impresa casearia (Parma)": '109',
    "Scienze, culture e politiche gastronomiche per il benessere (Roma - Sapienza)": '111',
    "Scienze e culture enogastronomiche (Roma - Tre)": '113',
    "Scienze e Tecnologie Agrarie (Reggio Calabria)": '115', # codice 5
    "Scienze e Tecnologie Alimentari (Reggio Calabria)": '116', # codice 6
    "Scienze Forestali e Ambientali (Reggio Calabria)": '117', # codice 11
    "Gestione e valorizzazione delle risorse agrarie e delle aree protette (Salerno)": '119',
    "Viticoltura ed Enologia (Salento)": '121', # codice 22
    "Agribusiness (Siena)": '123',
    "Scienze agro-zootecniche (Sassari)": '125',
    "Scienze e Tecnologie Agrarie (Sassari)": '126', # codice 5
    "Scienze Forestali e Ambientali (Sassari)": '127',
    "Tecnologie viticole, enologiche, alimentari (Sassari)": '128',
    "Intensificazione sostenibile delle produzioni ortofrutticole di qualità (Teramo)": '130',
    "Scienze e culture gastronomiche per la sostenibilità (Teramo)": '131',
    "Scienze e Tecnologie Alimentari (Teramo)": '132',
    "Tutela e benessere animale (Teramo)": '133',
    "Viticoltura ed Enologia (Teramo)": '134', # codice 22
    "Viticoltura ed Enologia (Trento)": '136', # codice 22
    "Scienze e Tecnologie Agrarie (Torino)": '138', # codice 5
    "Scienze e Tecnologie per la Montagna (Torino)": '139',
    "Scienze Forestali e Ambientali (Torino)": '140', # codice 11
    "Sistemi Zootecnici Sostenibili (Torino)": '141',
    "Tecniche di Assistenza Veterinaria (Torino)": '142',
    "Tecnologie Alimentari (Torino)": '143',
    "Viticoltura ed Enologia (Torino)": '144', # codice 22
    "Gestione Sostenibile delle Foreste e del Verde Urbano (Tuscia)": '146',
    "Produzione sementiera e vivaismo (Tuscia)": '147',
    "Scienze Agrarie e Ambientali (Tuscia)": '148',
    "Scienze della Montagna (Tuscia)": '149',
    "Scienze Forestali e Ambientali (Tuscia)": '150', # codice 11
    "Scienze, Culture e Politiche Gastronomiche per il Benessere (Tuscia)": '151',
    "Tecnologie per gestione sostenibile dei sistemi zootecnici (Tuscia)": '152',
    "Tecnologie Alimentari e Enologiche (Tuscia)": '153',
    "Scienza e cultura del cibo (Udine)": '155',
    "Scienze Agrarie (Udine)": '156', # codice 10
    "Scienze e Tecnologie Alimentari (Udine)": '157', # codice 6
    "Tutela e salute animale (Udine)": '158',
    "Viticoltura ed Enologia (Udine)": '159', # codice 22
    "Innovazione e sostenibilità nella produzione industriale di alimenti (Verona)": '161',
    "Scienze e Tecnologie Viticole ed Enologiche (Verona)": '162'
}

# Università POT
pot_uni_dict = {
    "Università degli Studi di Bari Aldo Moro": '1',
    "Università degli Studi della Basilicata": '9',
    "Università degli Studi di Bologna": '13',
    "Università degli Studi di Brescia": '23',
    "Università degli Studi di Camerino": '25',
    "Università degli Studi della Campania Luigi Vanvitelli": '28',
    "Università degli Studi di Catania": '30',
    "Università degli Studi di Ferrara": '34',
    "Università degli Studi di Foggia": '36',
    "Università degli Studi di Firenze": '40',
    "Università degli Studi di Messina": '48',
    "Università degli Studi di Milano": '51',
    "Università degli Studi di Modena e Reggio Emilia": '61',
    "Università degli Studi del Molise": '63',
    "Università degli Studi di Napoli Federico II": '66',
    "Università degli Studi di Palermo": '72',
    "Università degli Studi di Padova": '80',
    "Università degli Studi di Perugia": '91',
    "Università di Pisa": '96',
    "Università Politecnica delle Marche": '100',
    "Università degli Studi di Parma": '105',
    "Università degli Studi Roma La Sapienza": '110',
    "Università degli Studi Roma Tre": '112',
    "Università degli Studi Mediterranea di Reggio Calabria": '114',
    "Università degli Studi di Salerno": '118',
    "Università del Salento": '120',
    "Università degli Studi di Siena": '122',
    "Università degli Studi di Sassari": '124',
    "Università degli Studi di Teramo": '129',
    "Università degli Studi di Trento": '135',
    "Università degli Studi di Torino": '137',
    "Università degli Studi della Tuscia": '145',
    "Università degli Studi di Udine": '154',
    "Università degli Studi di Verona": '160'
}

# Università - Corsi di laurea
uni_cdl_dict = {
    "Università degli studi di Bari Aldo Moro":[
        "Gestione degli spazi verdi, dei boschi e delle aree protette (Bari)",
        "Scienze Animali (Bari)",
        "Scienze delle Produzioni e delle Risorse del Mare (Bari)",
        "Scienze e Tecnologie Agrarie (Bari)",
        "Scienze e Tecnologie Alimentari (Bari)",
        "Scienze per la valorizzazione del patrimonio gastronomico (Bari)",
        "Tecniche per l'agricoltura sostenibile (Bari)"
    ],
    "Università degli Studi della Basilicata": [
        "Scienze Agrarie (Basilicata)",
        "Scienze Forestali e Ambientali (Basilicata)",
        "Tecnologie Alimentari (Basilicata)"
    ],
    "Università degli Studi di Bologna": [
        "Acquacoltura e igiene delle produzioni ittiche (Bologna)",
        "Economia e Marketing nel sistema agro-industriale (Bologna)",
        "Produzioni animali (Bologna)",
        "Scienze e cultura della gastronomia (Bologna)",
        "Scienze e tecnologie per il verde e il paesaggio (Bologna)",
        "Tecnologie agrarie (Bologna)",
        "Tecnologie alimentari (Bologna)",
        "Tecnologie per il territorio e l'ambiente agro-forestale (Bologna)",
        "Viticoltura ed enologia (Bologna)"
    ],
    "Università degli Studi di Brescia": [
        "Sistemi agricoli sostenibili (Brescia)"
    ],
    "Università degli Studi di Camerino": [
        "Scienze Gastronomiche (Camerino)",
        "Tecnico del benessere animale e delle produzioni (Camerino)"
    ],
    "Università degli Studi della Campania Luigi Vanvitelli": [
        "Scienze Agrarie e Forestali (Campania)"
    ],
    "Università degli Studi di Catania": [
        "Scienze e tecnologie agrarie (Catania)",
        "Scienze e tecnologie alimentari (Catania)",
        "Scienze e tecnologie per la ristorazione e distribuzione degli alimenti mediterranei (Catania)"
    ],
    "Università degli Studi di Ferrara": [
        "Tecnologie agrarie e acquacoltura del Delta (Ferrara)"
    ],
    "Università degli Studi di Foggia": [
        "Cultura e Sostenibilità dell'Enogastronomia (Foggia)",
        "Scienze e tecnologie agrarie (Foggia)",
        "Scienze e Tecnologie Alimentari (Foggia)"
    ],
    "Università degli Studi di Firenze": [
        "Scienze agrarie (Firenze)",
        "Scienze e Tecnologie per la Gestione degli Spazi Verdi e del Paesaggio (Firenze)",
        "Scienze Faunistiche (Firenze)",
        "Scienze Forestali e Ambientali (Firenze)",
        "Tecnologie alimentari (Firenze)",
        "Tecnologie e trasformazioni avanzate nel settore legno, arredo, edilizia (Firenze)",
        "Viticoltura ed Enologia (Firenze)"
    ],
    "Università degli Studi di Messina": [
        "Scienze tecnologie e sicurezza delle produzioni animali (Messina)",
        "Scienze e tecnologie agrarie per la transizione ecologica (Messina)",
    ],
    "Università degli Studi di Milano": [
        "Agricoltura Sostenibile (Milano)",
        "Allevamento e benessere degli animali d'affezione (Milano)",
        "Produzione e protezione delle piante e dei sistemi del verde (Milano)",
        "Scienze della ristorazione e distribuzione degli alimenti (Milano)",
        "Scienze delle produzioni animali (Milano)",
        "Scienze e Tecnologie per alimenti sostenibili (Milano)",
        "Sistemi digitali in agricoltura (Milano)",
        "Valorizzazione e tutela dell'ambiente e del territorio montano (Milano)",
        "Viticoltura e Enologia (Milano)"
    ],
    "Università degli Studi di Modena e Reggio Emilia": [
        "Scienze e Tecnologie agrarie e degli alimenti (Modena)",
    ],
    "Università degli Studi del Molise": [
        "Scienze e Tecnologie Agrarie e Forestali (Molise)",
        "Scienze e Tecnologie Alimentari (Molise)"
    ],
    "Università degli Studi di Napoli Federico II": [
        "Scienze Agrarie, Forestali e Ambientali (Napoli)",
        "Scienze Gastronomiche Mediterranee (Napoli)",
        "Tecnologie delle Produzioni animali (Vecchio Ordinamento) e Gestione degli Animali e delle Produzioni (Nuovo Ordinamento) (Napoli)",
        "Tecnologie alimentari (Napoli)",
        "Viticoltura ed enologia (Napoli)"
    ],
    "Università degli Studi di Palermo": [
        "Agroingegneria (Palermo)",
        "Scienze e Tecnologie Agrarie (Palermo)",
        "Scienze e Tecnologie Agroalimentari (Palermo)",
        "Sciene Forestali e Ambientali (Palermo)",
        "Scienze Gastronomiche (Palermo)",
        "Sistemi Agricoli Mediterranei (Palermo)",
        "Viticoltura ed Enologia (Palermo)"
    ],
    "Università degli Studi di Padova": [
        "Animal Care - Tutela del benessere animale (Padova)",
        "Produzioni Biologiche Vegetali (Padova)",
        "Scienze e Cultura della Gastronomia (Padova)",
        "Scienze e tecnologie agrarie (Padova)",
        "Scienze e Tecnologie Alimentari (Padova)",
        "Scienze e Tecnologie Animali (Padova)",
        "Scienze e Tecnologie Viticole ed Enologiche (Padova)",
        "Scienze Tecnico Assistenziali Veterinarie (Padova)",
        "Sicurezza Igienico-Sanitaria degli Alimenti (Padova)",
        "Tecnologie Forestali e Ambientali (Padova)"
    ],
    "Università degli Studi di Perugia": [
        "Scienze Agrarie e Ambientali (Perugia)",
        "Economia e cultura dell'alimentazione (Perugia)",
        "Produzioni animali (Perugia)",
        "Scienze e Tecnologie Agroalimentari (Perugia)"
    ],
    "Università di Pisa": [
        "Scienze Agrarie (Pisa)",
        "Viticoltura e Enologia (Pisa)",
        "Scienze e Tecnologie delle Produzioni Animali (Pisa)"
    ],
    "Università Politecnica delle Marche": [
        "Scienze e Tecnologie Agrarie (Marche)",
        "Scienze e Tecnologie Alimentari (Marche)",
        "Scienze Forestali ed Ambientali (Marche)",
        "Sistemi Agricoli Innovativi (Marche)"
    ],
    "Università degli Studi di Parma": [
        "Scienze e Tecnologie Alimentari (Parma)",
        "Scienze Gastronomiche (Parma)",
        "Scienze Zootecniche e Tecnologie delle Produzioni Animali (Parma)",
        "Tecnologie e gestione dell’impresa casearia (Parma)"
    ],
    "Università degli Studi Roma La Sapienza": [
        "Scienze, culture e politiche gastronomiche per il benessere (Roma - Sapienza)"
    ],
    "Università degli Studi Roma Tre": [
        "Scienze e culture enogastronomiche (Roma - Tre)"
    ],
    "Università degli Studi Mediterranea di Reggio Calabria": [
        "Scienze e Tecnologie Agrarie (Reggio Calabria)",
        "Scienze e Tecnologie Alimentari (Reggio Calabria)",
        "Scienze Forestali ed Ambientali (Reggio Calabria)"
    ],
    "Università degli Studi di Salerno": [
        "Gestione e valorizzazione delle risorse agrarie e delle aree protette (Salerno)"
    ],
    "Università del Salento": [
        "Viticoltura ed enologia (Salento)"
    ],
    "Università degli Studi di Siena": [
        "Agribusiness (Siena)"
    ],
    "Università degli Studi di Sassari": [
        "Scienze agro-zootecniche (Sassari)",
        "Scienze e tecnologie agrarie (Sassari)",
        "Scienze forestali e ambientali (Sassari)",
        "Tecnologie viticole, enologiche, alimentari (Sassari)"
    ],
    "Università degli Studi di Teramo": [
        "Intensificazione sostenibile delle produzioni ortofrutticole di qualità (Teramo)",
        "Scienze e culture gastronimiche per la sostenibilità (Teramo)",
        "Scienze e Tecnologie alimentari (Teramo)",
        "Tutela e benessere animale (Teramo)",
        "Viticoltura ed Enologia (Teramo)",
    ],
    "Università degli Studi di Trento": [
        "Viticoltura ed enologia (Trento)"
    ],
    "Università degli Studi di Torino": [
        "Scienze e Tecnologie Agrarie (Torino)",
        "Scienze e Tecnologie per la Montagna (Torino)",
        "Scienze Forestali e Ambientali (Torino)",
        "Sistemi Zootecnici Sostenibili (Torino)",
        "Tecniche di Assistenza Veterinaria (Torino)",
        "Tecnologie Alimentari (Torino)",
        "Viticoltura ed Enologia (Torino)"
    ],
    "Università degli Studi della Tuscia": [
        "Gestione Sostenibile delle Foreste e del Verde Urbano (Tuscia)",
        "Produzione sementiera e vivaismo (Tuscia)",
        "Scienze Agrarie ed Ambientali (Tuscia)",
        "Scienze della Montagna (Tuscia)",
        "Scienze Forestali e Ambientali (Tuscia)",
        "Scienze, Culture e Politiche Gastronomiche per il Benessere (Tuscia)",
        "Tecnologie per gestione sostenibile dei sistemi zootecnici (Tuscia)",
        "Tecnologie Alimentari e Enologiche (Tuscia)"
    ],
    "Università degli Studi di Udine": [
        "Scienza e cultura del cibo (Udine)",
        "Scienze Agrarie (Udine)",
        "Scienze e tecnologie alimentari (Udine)",
        "Tutela e salute animale (Udine)",
        "Viticoltura ed enologia (Udine)"
    ],
    "Università degli Studi di Verona": [
        "Innovazione e sostenibilità nella produzione industriale di alimenti (Verona)",
        "Scienze e tecnologie viticole ed enologiche (Verona)"
    ]
}

cdl_composed_dict = {
    "Tecnologie Alimentari" : ['9', '13', '40', '66'],
    "Scienze e Tecnologie Agrarie" : ['1', '36', '72', '80', '100', '114', '124', '137'],
    "Scienze Agrarie" : ['9', '96', '154'],
    "Scienze Forestali e Ambientali" : ['9', '40', '72', '114', '137', '145'],
    "Produzioni Animali" : ['13', '91'],
    "Viticoltura ed Enologia" : ['13', '66', '72', '96', '120', '129', '135', '137', '154'],
    "Scienze e Tecnologie Alimentari" : ['1', '36', '63', '80', '100', '105', '114', '154'],
    "Scienze Gastronomiche" : ['25', '72', '105'],
    "Scienze e Tecnologie Agroalimentari" : ['72', '91']
}

cdl_composed_code_dict = {
    "Tecnologie Alimentari" : '12',
    "Scienze e Tecnologie Agrarie" : '5',
    "Scienze Agrarie" : '10',
    "Scienze Forestali e Ambientali" : '11',
    "Produzioni Animali" : '16',
    "Viticoltura ed Enologia" : '22',
    "Scienze e Tecnologie Alimentari" : '6',
    "Scienze Gastronomiche" : '26',
    "Scienze e Tecnologie Agroalimentari" : '75'
}

# Dimensioni questionari
dimensions_dict_QPSS = {
    "Curiosità epistemica" : ['CURIOSITA_01', 'CURIOSITA_02', 'CURIOSITA_03', 'CURIOSITA_04', 'CURIOSITA_05'],
    "Organizzazione e\ngestione del tempo di studio" : ['ORG.TIME_01', 'ORG.TIME_02', 'ORG.TIME_03', 'ORG.TIME_04', 'ORG.TIME_05',
                                                        'ORG.TIME_06', 'ORG.TIME_07'],
    "Percezione della propria competenza" : ['QPCS_33', 'QPCS_35', 'QPCS_38', 'QPCS_49', 'QPCS_51', 'QPCS_52', 'QPCS_54', 'QPCS_55'],
    "Dare senso e prospettiva alla\npropria esistenza umana e lavorativa" : ['QPCS_34', 'QPCS_37', 'QPCS_43', 'QPCS_45', 'QPCS_50',
                                                                             'QPCS_53'],
    "Gestire processi riflessivi" : ['QPCS_05', 'QPCS_09', 'QPCS_10', 'QPCS_11', 'QPCS_12', 'QPCS_14', 'QPCS_16', 'QPCS_19',
                                      'QPCS_21', 'QPCS_28'],
    "Gestire se stessi\nnel lavoro e nell’apprendimento\n(autoregolazione e volizione)" : ['QPCS_04', 'QPCS_07', 'QPCS_08', 'QPCS_15', 'QPCS_22', 'QPCS_23',
                                                            'QPCS_25', 'QPCS_26', 'QPCS_31', 'QPCS_42'],
    "Gestire forme accentuate di ansietà" : ['QPCS_01', 'QPCS_02', 'QPCS_03', 'QPCS_06', 'QPCS_13', 'QPCS_17', 'QPCS_18',
                                              'QPCS_20', 'QPCS_24', 'QPCS_27', 'QPCS_29'],
    "Collaborare con altre\npersone nel lavoro e nell’apprendimento" : ['QPCS_30', 'QPCS_32', 'QPCS_36', 'QPCS_39', 
                                                                         'QPCS_40', 'QPCS_41', 'QPCS_44', 'QPCS_46', 
                                                                         'QPCS_47', 'QPCS_48']
}

dimensions_dict_QBEAP = {
    "Motivazione intrinseca": ['MOT_1', 'MOT_2', 'MOT_3', 'MOT_4'],
    "Motivazione identificata": ['MOT_5', 'MOT_6', 'MOT_7', 'MOT_8'],
    "Motivazione estrinseca": ['MOT_9', 'MOT_10', 'MOT_11', 'MOT_12'],
    "Autoefficacia accademica": ['AUTOEFFICACIA_1', 'AUTOEFFICACIA_2', 'AUTOEFFICACIA_3',
                                 'AUTOEFFICACIA_4', 'AUTOEFFICACIA_5'],
    "Mentalità di crescita": ['GR_MIND_1', 'GR_MIND_2', 'GR_MIND_3', 'GR_MIND_4',
                              'GR_MIND_5', 'GR_MIND_6', 'GR_MIND_7', 'GR_MIND_8'],
    "Consapevolezza del proprio apprendimento": ['SELF_LEARN_1', 'SELF_LEARN_2', 'SELF_LEARN_3',
                                                 'SELF_LEARN_4', 'SELF_LEARN_5', 'SELF_LEARN_6', 'SELF_LEARN_7'],
    "Rispetto delle scadenze": ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4'],
    "Interazione con docenti": ['IIS_1', 'IIS_2', 'IIS_3', 'IIS_4'],
    "Interesse dei docenti per lo sviluppo degli studenti e dell'insegnamento": ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_8'],
    "Interazioni tra pari": ['IIS_9', 'IIS_10', 'IIS_11', 'IIS_12', 'IIS_13'],
    "Sviluppo accademico e intellettuale": ['IIS_14', 'IIS_15', 'IIS_16', 'IIS_17', 'IIS_18'],
    "Impegno verso gli obiettivi universitari": ['IIS_19', 'IIS_20', 'IIS_21', 'IIS_22'],
    "Ripensamenti sul percorso universitario": ['INT_ABB_1', 'INT_ABB_2', 'INT_ABB_3', 'INT_ABB_4'],
    "Consapevolezza dei propri interessi professionali": ['RIS_CARR_7', 'RIS_CARR_8', 'RIS_CARR_9'],
    "Chiarezza degli obiettivi professionali": ['RIS_CARR_1', 'RIS_CARR_2', 'RIS_CARR_3'],
    "Ricerca di informazioni sul futuro professionale": ['RIS_CARR_4', 'RIS_CARR_5', 'RIS_CARR_6'],
    "Benessere percepito": ['WEMWBS_1', 'WEMWBS_2', 'WEMWBS_3', 'WEMWBS_4',
                            'WEMWBS_5', 'WEMWBS_6', 'WEMWBS_7', 'WEMWBS_8',
                            'WEMWBS_9', 'WEMWBS_10', 'WEMWBS_11', 'WEMWBS_12']
}

dimensions_dict_riflessioni = {
    "per tutti gli studenti e le studentesse indipendentemente\ndai risultati ottenuti agli esami universitari" : ['LAB_UTILITA01'],
    "per affrontare positivamente lo studio universitario" : ['LAB_UTILITA02'],
    "per tutti gli studenti e le studentesse indipendentemente\ndai risultati ottenuti alle superiori" : ['LAB_UTILITA03'],
    "per migliorare i propri risultati universitari" : ['LAB_UTILITA04'],
    "se svolto all'inizio del percorso\nuniversitario (primo semestre del primo anno)" : ['LAB_TEMPO02']
}

dimensions_dict_POT = {
    "Motivazione intrinseca": ['MOT_1', 'MOT_2', 'MOT_3', 'MOT_4'],
    "Motivazione identificata": ['MOT_5', 'MOT_6', 'MOT_7', 'MOT_8'],
    "Motivazione estrinseca": ['MOT_9', 'MOT_10', 'MOT_11', 'MOT_12'],
    "Autoefficacia accademica": ['AUTOEFFICACIA_1', 'AUTOEFFICACIA_2', 'AUTOEFFICACIA_3',
                                 'AUTOEFFICACIA_4', 'AUTOEFFICACIA_5'],
    "Consapevolezza del proprio apprendimento": ['SELF_LEARN_1', 'SELF_LEARN_2', 'SELF_LEARN_3',
                                                 'SELF_LEARN_4', 'SELF_LEARN_5', 'SELF_LEARN_6', 'SELF_LEARN_7'],
    "Rispetto delle scadenze": ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4'],
    "Benessere percepito": ['WHO_1', 'WHO_2', 'WHO_3', 'WHO_4', 'WHO_5'],
    "Interesse dei docenti riguardo lo\nsviluppo dellə studentə e dell'insegnamento": ['IIS_1', 'IIS_2', 'IIS_3', 'IIS_4'],
    "Interazioni tra pari": ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_8', 'IIS_9'],
    "Impegno verso gli obiettivi universitari": ['IIS_10', 'IIS_11', 'IIS_12', 'IIS_13'],
    "Consapevolezza dei propri\ninteressi professionali": ['RIS_CARR_7', 'RIS_CARR_8', 'RIS_CARR_9'],
    "Chiarezza degli obiettivi professionali": ['RIS_CARR_1', 'RIS_CARR_2', 'RIS_CARR_3'],
    "Ricerca di informazioni\nsul benessere professionale": ['RIS_CARR_4', 'RIS_CARR_5', 'RIS_CARR_6'],
    "Ripensamenti sul percorso universitario": ['INT_ABB_1', 'INT_ABB_2', 'INT_ABB_3', 'INT_ABB_4'],
    "Genitori": ['HELICOPTER_1', 'HELICOPTER_2', 'HELICOPTER_3', 'HELICOPTER_4', 'HELICOPTER_5', 'HELICOPTER_6',
                 'HELICOPTER_7', 'HELICOPTER_8', 'HELICOPTER_9', 'HELICOPTER_10', 'HELICOPTER_11',
                 'HELICOPTER_12', 'HELICOPTER_13', 'HELICOPTER_14', 'HELICOPTER_15']
}

############################################################################################################
# Class
class PDF(FPDF):
    pdf_w=210
    pdf_h=297
    
    def add_fonts(self):
        self.add_font('OSr', 'R', 'fonts' + os.sep + 'OpenSans-Regular.ttf', uni=True)
        self.add_font('OSb', 'B', 'fonts' + os.sep + 'OpenSans-Bold.ttf', uni=True)
        self.add_font('OSi', 'I', 'fonts' + os.sep + 'OpenSans-Italic.ttf', uni=True)

    def bg(self, page_name, dir):        
        # Set the size of the page to match the existing PDF file
        self.set_auto_page_break(auto=True, margin=0)

        # Import the existing PDF as an image onto the blank page
        self.image('pages' + os.sep + dir + os.sep + page_name + '.png', x=0, y=0, w=self.w, h=self.h)

############################################################################################################
# Exported functions

def get_exported_data_cdls_1(file_csv, cdl):
    df = pd.read_csv(file_csv, sep=';', dtype=object, encoding = 'ISO-8859-1')
    labs = {}

    file1 = openpyxl.load_workbook('files'+os.sep+'lab.xlsx', data_only=True)
    labs = file1['Laboratori e colloqui']

    for row in labs.iter_rows(min_row=3, max_row=119, min_col=1, max_col=1, values_only=True):
        if row[0] == int(cdls_dict[cdl]):
            n_org_tempo = labs.cell(row=row[0]+2, column=7).value if labs.cell(row=row[0]+2, column=7).value else 0
            n_strat_appr = labs.cell(row=row[0]+2, column=8).value if labs.cell(row=row[0]+2, column=8).value else 0
            n_gest_esame = labs.cell(row=row[0]+2, column=9).value if labs.cell(row=row[0]+2, column=9).value else 0
            n_colloqui = labs.cell(row=row[0]+2, column=3).value if labs.cell(row=row[0]+2, column=3).value else 0
    
    df = df.loc[(df['Progress'] == '100')]
    df_1 = df.loc[(df['Anno'] == '1') & (df['CdS'] == cdls_dict[cdl])]
    # print(df_1)

    data = {
        'cdl': cdl,
        'cdl_code': cdls_dict[cdl],
        'numero_studenti_1': df_1.shape[0],
        'numero_maschi_1': df_1.loc[df_1['GENERE'] == '1'].shape[0],
        'numero_femmine_1': df_1.loc[df_1['GENERE'] == '2'].shape[0],
        'numero_nonbin_1': df_1.loc[df_1['GENERE'] == '3'].shape[0],
        'numero_altro_1': df_1.loc[df_1['GENERE'] == '4'].shape[0],
        'numero_eta_media_1': round(pd.to_numeric(df_1['ETA'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_1': round(pd.to_numeric(df_1['ETA'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_1': df_1.loc[df_1['FREQ'] == '1'].shape[0],
        'no_partecip_1': df_1.loc[df_1['M1_NO'] == '1'].shape[0],

        'NUMLAB_1': n_org_tempo + n_strat_appr + n_gest_esame,
        'NUMCOLLOQUI_1': n_colloqui

    }

    if np.isnan(data['numero_eta_media_1']):
        data['numero_eta_media_1'] = 0.0

    if np.isnan(data['deviazione_standard_1']):
        data['deviazione_standard_1'] = 0.0

    for dim in dimensions_dict_QPSS:
        # print("Dimensione:", dim)
        colonne_dimensione = dimensions_dict_QPSS[dim]
        minimo_teorico = 1 * len(colonne_dimensione)
        massimo_teorico = 4 * len(colonne_dimensione)

        punti_t_per_riga = []
        punti_t_min_40 = []
        punti_t_40_60 = []
        punti_t_magg_60 = []
        punti_z_per_riga = []
        punteggi_normalizzati_pop = []

        for index, row in df.iterrows():
            
            punteggi_grezzi_scala_pop = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                if pd.notna(row[colonna]):
                    if dim == "Gestire forme accentuate di ansietà":
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_pop.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_pop.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_pop.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_pop.append(1)
                    else:
                        punteggi_grezzi_scala_pop.append(int(row[colonna]))
                        
            punteggio_grezzo_totale_pop = sum(punteggi_grezzi_scala_pop)
            punteggio_normalizzato_pop = (punteggio_grezzo_totale_pop - minimo_teorico) / (massimo_teorico - minimo_teorico)
        
            punteggi_normalizzati_pop.append(punteggio_normalizzato_pop)
    
        # Itera attraverso le righe del dataframe
        for index, row in df_1.iterrows():
            
            punteggi_grezzi_scala = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if dim == "Gestire forme accentuate di ansietà":
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala.append(1)
                    else:
                        punteggi_grezzi_scala.append(int(row[colonna]))
                        
            punteggio_grezzo_totale = sum(punteggi_grezzi_scala)
            punteggio_normalizzato = (punteggio_grezzo_totale - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga.append(punteggio_normalizzato)  
            punti_t_per_riga.append(punteggio_normalizzato)
        
        media_punteggi_normalizzati_pop = np.mean(punteggi_normalizzati_pop)
        deviazione_standard_punteggi_normalizzati_pop = np.std(punteggi_normalizzati_pop)

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga[i] = 0 #!!!
            else:
                punti_z_per_riga[i] = ((punti_t_per_riga[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga[i] = 50 + (10 * punti_z_per_riga[i])
            if punti_t_per_riga[i] < 40:
                punti_t_min_40.append(punti_t_per_riga[i])
            elif punti_t_per_riga[i] >= 40 and punti_t_per_riga[i] <= 60:
                punti_t_40_60.append(punti_t_per_riga[i])
            else:
                punti_t_magg_60.append(punti_t_per_riga[i])
    
        data[f'{dim}_min_40_1'] = punti_t_min_40
        data[f'{dim}_40_60_1'] = punti_t_40_60
        data[f'{dim}_magg_60_1'] = punti_t_magg_60

    for dim in dimensions_dict_riflessioni:
        colonne_dimensione = dimensions_dict_riflessioni[dim]
        punteggi_per_nulla_daccordo = []
        punteggi_solo_in_parte_daccordo = []
        punteggi_abbastanza_daccordo = []
        punteggi_pienamente_daccordo = []

        for index, row in df_1.iterrows():
            for colonna in colonne_dimensione:
                if pd.notna(row[colonna]):
                    if (row[colonna] == "1"):
                        punteggi_per_nulla_daccordo.append(1)
                    elif (row[colonna] == "2"):
                        punteggi_solo_in_parte_daccordo.append(1)
                    elif (row[colonna] == "3"):
                        punteggi_abbastanza_daccordo.append(1)
                    elif (row[colonna] == "4"):
                        punteggi_pienamente_daccordo.append(1)
        
        data[f'{dim}_per_nulla_daccordo_1'] = len(punteggi_per_nulla_daccordo)
        data[f'{dim}_solo_in_parte_daccordo_1'] = len(punteggi_solo_in_parte_daccordo)
        data[f'{dim}_abbastanza_daccordo_1'] = len(punteggi_abbastanza_daccordo)
        data[f'{dim}_pienamente_daccordo_1'] = len(punteggi_pienamente_daccordo)

    return data

def get_exported_data_cdls_23(file_csv, cdl):
    df = pd.read_csv(file_csv, sep=';', dtype=object, encoding = 'ISO-8859-1')
    labs = {}

    file1 = openpyxl.load_workbook('files'+os.sep+'lab.xlsx', data_only=True)
    labs = file1['Laboratori e colloqui']

    # Funzione per pulire il valore della cella
    def clean_cell_value(value):
        if isinstance(value, str):
            value = value.replace('+', '').replace('\xa0', '').strip()
        return int(value) if value else 0

    # Itera sulle righe del foglio di lavoro
    for row in labs.iter_rows(min_row=3, max_row=119, min_col=1, max_col=1, values_only=True):
        if row[0] == int(cdls_dict[cdl]):
            n_org_tempo = clean_cell_value(labs.cell(row=row[0]+2, column=10).value or 0) + \
                        clean_cell_value(labs.cell(row=row[0]+2, column=17).value or 0) + \
                        clean_cell_value(labs.cell(row=row[0]+2, column=18).value or 0)
        
            n_gest_emo = clean_cell_value(labs.cell(row=row[0]+2, column=11).value)
            n_pren_dec = clean_cell_value(labs.cell(row=row[0]+2, column=12).value or 0) + \
                        clean_cell_value(labs.cell(row=row[0]+2, column=16).value or 0)
            n_facc_chiar = clean_cell_value(labs.cell(row=row[0]+2, column=13).value)
            n_scelta_lm = clean_cell_value(labs.cell(row=row[0]+2, column=14).value)
            n_personal_skills = clean_cell_value(labs.cell(row=row[0]+2, column=15).value)
            n_colloqui_23 = clean_cell_value(labs.cell(row=row[0]+2, column=5).value)
    
    df = df.loc[(df['Progress'] == '100')]
    df_2 = df.loc[(df['ANNO'] == '2') & (df['CdS'] == cdls_dict[cdl])]
    df_3 = df.loc[(df['ANNO'] == '3') & (df['CdS'] == cdls_dict[cdl])]
    df_fc = df.loc[(df['ANNO'].isin(['4', '5', '6'])) & (df['CdS'] == cdls_dict[cdl])]
    df_4 = df.loc[(df['ANNO'] == '4') & (df['CdS'] == cdls_dict[cdl])]
    df_5 = df.loc[(df['ANNO'] == '5') & (df['CdS'] == cdls_dict[cdl])]
    df_f2 = df.loc[(df['ANNO'].isin(['5', '6'])) & (df['CdS'] == cdls_dict[cdl])]
    df_f3 = df.loc[(df['ANNO'] == '6') & (df['CdS'] == cdls_dict[cdl])]

    df_3fc = df.loc[(df['ANNO'].isin(['3', '4', '5', '6'])) & (df['CdS'] == cdls_dict[cdl])]
    df_4f2 = df.loc[(df['ANNO'].isin(['4', '5', '6'])) & (df['CdS'] == cdls_dict[cdl])]
    df_5f3 = df.loc[(df['ANNO'].isin(['5', '6'])) & (df['CdS'] == cdls_dict[cdl])]
    # print(df_fc)

    data = {
        'cdl': cdl,
        'cdl_code': cdls_dict[cdl],
        'numero_studenti_2': df_2.shape[0],
        'numero_maschi_2': df_2.loc[df_2['GENERE'] == '1'].shape[0],
        'numero_femmine_2': df_2.loc[df_2['GENERE'] == '2'].shape[0],
        'numero_nonbin_2': df_2.loc[df_2['GENERE'] == '3'].shape[0],
        'numero_altro_2': df_2.loc[df_2['GENERE'] == '4'].shape[0],
        'numero_eta_media_2': round(pd.to_numeric(df_2['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_2': round(pd.to_numeric(df_2['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_2': df_2.loc[df_2['FREQ'] == '1'].shape[0],

        'numero_studenti_3': df_3.shape[0],
        'numero_maschi_3': df_3.loc[df_3['GENERE'] == '1'].shape[0],
        'numero_femmine_3': df_3.loc[df_3['GENERE'] == '2'].shape[0],
        'numero_nonbin_3': df_3.loc[df_3['GENERE'] == '3'].shape[0],
        'numero_altro_3': df_3.loc[df_3['GENERE'] == '4'].shape[0],
        'numero_eta_media_3': round(pd.to_numeric(df_3['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_3': round(pd.to_numeric(df_3['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_3': df_3.loc[df_3['FREQ'] == '1'].shape[0],

        'numero_studenti_fc': df_fc.shape[0],
        'numero_maschi_fc': df_fc.loc[df_fc['GENERE'] == '1'].shape[0],
        'numero_femmine_fc': df_fc.loc[df_fc['GENERE'] == '2'].shape[0],
        'numero_nonbin_fc': df_fc.loc[df_fc['GENERE'] == '3'].shape[0],
        'numero_altro_fc': df_fc.loc[df_fc['GENERE'] == '4'].shape[0],
        'numero_eta_media_fc': round(pd.to_numeric(df_fc['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_fc': round(pd.to_numeric(df_fc['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_fc': df_fc.loc[df_fc['FREQ'] == '1'].shape[0],

        'numero_studenti_4': df_4.shape[0],
        'numero_maschi_4': df_4.loc[df_4['GENERE'] == '1'].shape[0],
        'numero_femmine_4': df_4.loc[df_4['GENERE'] == '2'].shape[0],
        'numero_nonbin_4': df_4.loc[df_4['GENERE'] == '3'].shape[0],
        'numero_altro_4': df_4.loc[df_4['GENERE'] == '4'].shape[0],
        'numero_eta_media_4': round(pd.to_numeric(df_4['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_4': round(pd.to_numeric(df_4['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_4': df_4.loc[df_4['FREQ'] == '1'].shape[0],

        'numero_studenti_5': df_5.shape[0],
        'numero_maschi_5': df_5.loc[df_5['GENERE'] == '1'].shape[0],
        'numero_femmine_5': df_5.loc[df_5['GENERE'] == '2'].shape[0],
        'numero_nonbin_5': df_5.loc[df_5['GENERE'] == '3'].shape[0],
        'numero_altro_5': df_5.loc[df_5['GENERE'] == '4'].shape[0],
        'numero_eta_media_5': round(pd.to_numeric(df_5['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_5': round(pd.to_numeric(df_5['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_5': df_5.loc[df_5['FREQ'] == '1'].shape[0],

        'numero_studenti_f2': df_f2.shape[0],
        'numero_maschi_f2': df_f2.loc[df_f2['GENERE'] == '1'].shape[0],
        'numero_femmine_f2': df_f2.loc[df_f2['GENERE'] == '2'].shape[0],
        'numero_nonbin_f2': df_f2.loc[df_f2['GENERE'] == '3'].shape[0],
        'numero_altro_f2': df_f2.loc[df_f2['GENERE'] == '4'].shape[0],
        'numero_eta_media_f2': round(pd.to_numeric(df_f2['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_f2': round(pd.to_numeric(df_f2['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_f2': df_f2.loc[df_f2['FREQ'] == '1'].shape[0],

        'numero_studenti_f3': df_f3.shape[0],
        'numero_maschi_f3': df_f3.loc[df_f3['GENERE'] == '1'].shape[0],
        'numero_femmine_f3': df_f3.loc[df_f3['GENERE'] == '2'].shape[0],
        'numero_nonbin_f3': df_f3.loc[df_f3['GENERE'] == '3'].shape[0],
        'numero_altro_f3': df_f3.loc[df_f3['GENERE'] == '4'].shape[0],
        'numero_eta_media_f3': round(pd.to_numeric(df_f3['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_f3': round(pd.to_numeric(df_f3['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_f3': df_f3.loc[df_f3['FREQ'] == '1'].shape[0],

        'numero_eta_media_3fc': round(pd.to_numeric(df_3fc['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'numero_eta_media_4f2': round(pd.to_numeric(df_4f2['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'numero_eta_media_5f3': round(pd.to_numeric(df_5f3['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        
        'NUMLAB_23': n_org_tempo + n_gest_emo + n_pren_dec + n_facc_chiar + n_scelta_lm + n_personal_skills,
        'NUMCOLLOQUI_23': n_colloqui_23

    }

    if np.isnan(data['numero_eta_media_2']):
        data['numero_eta_media_2'] = 0.0

    if np.isnan(data['deviazione_standard_2']):
        data['deviazione_standard_2'] = 0.0
    
    if np.isnan(data['numero_eta_media_3']):
        data['numero_eta_media_3'] = 0.0

    if np.isnan(data['deviazione_standard_3']):
        data['deviazione_standard_3'] = 0.0

    if np.isnan(data['numero_eta_media_fc']):
        data['numero_eta_media_fc'] = 0.0

    if np.isnan(data['deviazione_standard_fc']):
        data['deviazione_standard_fc'] = 0.0

    if np.isnan(data['numero_eta_media_4']):
        data['numero_eta_media_4'] = 0.0

    if np.isnan(data['deviazione_standard_4']):
        data['deviazione_standard_4'] = 0.0

    if np.isnan(data['numero_eta_media_5']):
        data['numero_eta_media_5'] = 0.0

    if np.isnan(data['deviazione_standard_5']):
        data['deviazione_standard_5'] = 0.0

    if np.isnan(data['numero_eta_media_f2']):
        data['numero_eta_media_f2'] = 0.0

    if np.isnan(data['deviazione_standard_f2']):
        data['deviazione_standard_f2'] = 0.0

    if np.isnan(data['numero_eta_media_f3']):
        data['numero_eta_media_f3'] = 0.0

    if np.isnan(data['deviazione_standard_f3']):
        data['deviazione_standard_f3'] = 0.0

    if np.isnan(data['numero_eta_media_3fc']):
        data['numero_eta_media_3fc'] = 0.0

    if np.isnan(data['numero_eta_media_4f2']):
        data['numero_eta_media_4f2'] = 0.0

    if np.isnan(data['numero_eta_media_5f3']):
        data['numero_eta_media_5f3'] = 0.0

    for dim in dimensions_dict_QBEAP:
        # print("Dimensione:", dim)
        colonne_dimensione = dimensions_dict_QBEAP[dim]
        minimo_teorico = 1 * len(colonne_dimensione)
        massimo_teorico = 5 * len(colonne_dimensione)
        punteggi_normalizzati_pop = []

        # Punti T e Z secondo anno
        punti_t_per_riga_2 = []
        punti_t_min_40_2 = []
        punti_t_40_60_2 = []
        punti_t_magg_60_2 = []
        punti_z_per_riga_2 = []

        # Punti T e Z terzo anno
        punti_t_per_riga_3 = []
        punti_t_min_40_3 = []
        punti_t_40_60_3 = []
        punti_t_magg_60_3 = []
        punti_z_per_riga_3 = []

        # Punti T e Z fuori corso
        punti_t_per_riga_fc = []
        punti_t_min_40_fc = []
        punti_t_40_60_fc = []
        punti_t_magg_60_fc = []
        punti_z_per_riga_fc = []

        # Punti T e Z quarto anno
        punti_t_per_riga_4 = []
        punti_t_min_40_4 = []
        punti_t_40_60_4 = []
        punti_t_magg_60_4 = []
        punti_z_per_riga_4 = []

        # Punti T e Z quinto anno
        punti_t_per_riga_5 = []
        punti_t_min_40_5 = []
        punti_t_40_60_5 = []
        punti_t_magg_60_5 = []
        punti_z_per_riga_5 = []

        # Punti T e Z fuori corso cdl 1 2 4
        punti_t_per_riga_f2 = []
        punti_t_min_40_f2 = []
        punti_t_40_60_f2 = []
        punti_t_magg_60_f2 = []
        punti_z_per_riga_f2 = []

        # Punti T e Z fuori corso cdl 1 3 5
        punti_t_per_riga_f3 = []
        punti_t_min_40_f3 = []
        punti_t_40_60_f3 = []
        punti_t_magg_60_f3 = []
        punti_z_per_riga_f3 = []
    
        # Itera attraverso le righe del dataframe
        for index, row in df.iterrows():
            punteggi_grezzi_scala_pop = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_pop.append(5)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_pop.append(4)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_pop.append(3)    
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_pop.append(2)
                        elif (row[colonna] == "5"):
                            punteggi_grezzi_scala_pop.append(1)
                    else:
                        if row[colonna].strip().isdigit():
                            punteggi_grezzi_scala_pop.append(int(row[colonna]))
                        else:
                            punteggi_grezzi_scala_pop.append(0)
                        
            punteggio_grezzo_totale_pop = sum(punteggi_grezzi_scala_pop)
            punteggio_normalizzato_pop = (punteggio_grezzo_totale_pop - minimo_teorico) / (massimo_teorico - minimo_teorico)

            punteggi_normalizzati_pop.append(punteggio_normalizzato_pop)

        # Itera attraverso le righe del dataframe
        for index, row in df_2.iterrows():
            
            punteggi_grezzi_scala_2 = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_2.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_2.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_2.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_2.append(1)
                    else:
                        if row[colonna].strip().isdigit():
                            punteggi_grezzi_scala_2.append(int(row[colonna]))
                        else:
                            punteggi_grezzi_scala_2.append(0)
                        
            punteggio_grezzo_totale_2 = sum(punteggi_grezzi_scala_2)
            punteggio_normalizzato_2 = (punteggio_grezzo_totale_2 - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga_2.append(punteggio_normalizzato_2)  
            punti_t_per_riga_2.append(punteggio_normalizzato_2)

        # Itera attraverso le righe del dataframe
        for index, row in df_3.iterrows():
            
            punteggi_grezzi_scala_3 = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_3.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_3.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_3.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_3.append(1)
                    else:
                        if row[colonna].strip().isdigit():
                            punteggi_grezzi_scala_3.append(int(row[colonna]))
                        else:
                            punteggi_grezzi_scala_3.append(0)
                        
            punteggio_grezzo_totale_3 = sum(punteggi_grezzi_scala_3)
            punteggio_normalizzato_3 = (punteggio_grezzo_totale_3 - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga_3.append(punteggio_normalizzato_3)  
            punti_t_per_riga_3.append(punteggio_normalizzato_3)

        # Itera attraverso le righe del dataframe
        for index, row in df_fc.iterrows():
                
                punteggi_grezzi_scala_fc = []
    
                # Calcola il punteggio grezzo totale della scala
                for colonna in colonne_dimensione:
                    # print("Colonna:", colonna)
                    if pd.notna(row[colonna]):
                        if dim == (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                            if (row[colonna] == "1"):
                                punteggi_grezzi_scala_fc.append(4)
                            elif (row[colonna] == "2"):
                                punteggi_grezzi_scala_fc.append(3)
                            elif (row[colonna] == "3"):
                                punteggi_grezzi_scala_fc.append(2)
                            elif (row[colonna] == "4"):
                                punteggi_grezzi_scala_fc.append(1)
                        else:
                            if row[colonna].strip().isdigit():
                                punteggi_grezzi_scala_fc.append(int(row[colonna]))
                            else:
                                punteggi_grezzi_scala_fc.append(0)
                            
                punteggio_grezzo_totale_fc = sum(punteggi_grezzi_scala_fc)
                punteggio_normalizzato_fc = (punteggio_grezzo_totale_fc - minimo_teorico) / (massimo_teorico - minimo_teorico)      
                punti_z_per_riga_fc.append(punteggio_normalizzato_fc)  
                punti_t_per_riga_fc.append(punteggio_normalizzato_fc)

        # Itera attraverso le righe del dataframe
        for index, row in df_4.iterrows():
                
                punteggi_grezzi_scala_4 = []
    
                # Calcola il punteggio grezzo totale della scala
                for colonna in colonne_dimensione:
                    # print("Colonna:", colonna)
                    if pd.notna(row[colonna]):
                        if dim == (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                            if (row[colonna] == "1"):
                                punteggi_grezzi_scala_4.append(4)
                            elif (row[colonna] == "2"):
                                punteggi_grezzi_scala_4.append(3)
                            elif (row[colonna] == "3"):
                                punteggi_grezzi_scala_4.append(2)
                            elif (row[colonna] == "4"):
                                punteggi_grezzi_scala_4.append(1)
                        else:
                            if row[colonna].strip().isdigit():
                                punteggi_grezzi_scala_4.append(int(row[colonna]))
                            else:
                                punteggi_grezzi_scala_4.append(0)
                            
                punteggio_grezzo_totale_4 = sum(punteggi_grezzi_scala_4)
                punteggio_normalizzato_4 = (punteggio_grezzo_totale_4 - minimo_teorico) / (massimo_teorico - minimo_teorico)      
                punti_z_per_riga_4.append(punteggio_normalizzato_4)  
                punti_t_per_riga_4.append(punteggio_normalizzato_4)

        # Itera attraverso le righe del dataframe
        for index, row in df_5.iterrows():

                punteggi_grezzi_scala_5 = []

                # Calcola il punteggio grezzo totale della scala
                for colonna in colonne_dimensione:
                    # print("Colonna:", colonna)
                    if pd.notna(row[colonna]):
                        if dim == (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                            if (row[colonna] == "1"):
                                punteggi_grezzi_scala_5.append(4)
                            elif (row[colonna] == "2"):
                                punteggi_grezzi_scala_5.append(3)
                            elif (row[colonna] == "3"):
                                punteggi_grezzi_scala_5.append(2)
                            elif (row[colonna] == "4"):
                                punteggi_grezzi_scala_5.append(1)
                        else:
                            if row[colonna].strip().isdigit():
                                punteggi_grezzi_scala_5.append(int(row[colonna]))
                            else:
                                punteggi_grezzi_scala_5.append(0)

                punteggio_grezzo_totale_5 = sum(punteggi_grezzi_scala_5)
                punteggio_normalizzato_5 = (punteggio_grezzo_totale_5 - minimo_teorico) / (massimo_teorico - minimo_teorico)
                punti_z_per_riga_5.append(punteggio_normalizzato_5)
                punti_t_per_riga_5.append(punteggio_normalizzato_5)

        # Itera attraverso le righe del dataframe
        for index, row in df_f2.iterrows():

                punteggi_grezzi_scala_f2 = []

                # Calcola il punteggio grezzo totale della scala
                for colonna in colonne_dimensione:
                    # print("Colonna:", colonna)
                    if pd.notna(row[colonna]):
                        if dim == (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                            if (row[colonna] == "1"):
                                punteggi_grezzi_scala_f2.append(4)
                            elif (row[colonna] == "2"):
                                punteggi_grezzi_scala_f2.append(3)
                            elif (row[colonna] == "3"):
                                punteggi_grezzi_scala_f2.append(2)
                            elif (row[colonna] == "4"):
                                punteggi_grezzi_scala_f2.append(1)
                        else:
                            if row[colonna].strip().isdigit():
                                punteggi_grezzi_scala_f2.append(int(row[colonna]))
                            else:
                                punteggi_grezzi_scala_f2.append(0)

                punteggio_grezzo_totale_f2 = sum(punteggi_grezzi_scala_f2)
                punteggio_normalizzato_f2 = (punteggio_grezzo_totale_f2 - minimo_teorico) / (massimo_teorico - minimo_teorico)
                punti_z_per_riga_f2.append(punteggio_normalizzato_f2)
                punti_t_per_riga_f2.append(punteggio_normalizzato_f2)

        # Itera attraverso le righe del dataframe
        for index, row in df_f3.iterrows():
                
                    punteggi_grezzi_scala_f3 = []
    
                    # Calcola il punteggio grezzo totale della scala
                    for colonna in colonne_dimensione:
                        # print("Colonna:", colonna)
                        if pd.notna(row[colonna]):
                            if dim == (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                            or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                            or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                                if (row[colonna] == "1"):
                                    punteggi_grezzi_scala_f3.append(4)
                                elif (row[colonna] == "2"):
                                    punteggi_grezzi_scala_f3.append(3)
                                elif (row[colonna] == "3"):
                                    punteggi_grezzi_scala_f3.append(2)
                                elif (row[colonna] == "4"):
                                    punteggi_grezzi_scala_f3.append(1)
                            else:
                                if row[colonna].strip().isdigit():
                                    punteggi_grezzi_scala_f3.append(int(row[colonna]))
                                else:
                                    punteggi_grezzi_scala_f3.append(0)
    
                    punteggio_grezzo_totale_f3 = sum(punteggi_grezzi_scala_f3)
                    punteggio_normalizzato_f3 = (punteggio_grezzo_totale_f3 - minimo_teorico) / (massimo_teorico - minimo_teorico)
                    punti_z_per_riga_f3.append(punteggio_normalizzato_f3)
        
        
        # Calcola il punteggio z utilizzando media e deviazione standard della popolazione (Ateneo)
        media_punteggi_normalizzati_pop = np.mean(punteggi_normalizzati_pop)
        # print("Media punteggi normalizzati popolazione:", media_punteggi_normalizzati_pop)
        deviazione_standard_punteggi_normalizzati_pop = np.std(punteggi_normalizzati_pop)
        # print("Deviazione standard punteggi normalizzati popolazione:", deviazione_standard_punteggi_normalizzati_pop)

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_2)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_2[i] = 0 # !!!
            else:
                punti_z_per_riga_2[i] = ((punti_t_per_riga_2[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)           
            punti_t_per_riga_2[i] = 50 + (10 * punti_z_per_riga_2[i])
            if punti_t_per_riga_2[i] < 40:
                punti_t_min_40_2.append(punti_t_per_riga_2[i])
            elif punti_t_per_riga_2[i] >= 40 and punti_t_per_riga_2[i] <= 60:
                punti_t_40_60_2.append(punti_t_per_riga_2[i])
            else:
                punti_t_magg_60_2.append(punti_t_per_riga_2[i])          

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_3)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_3[i] = 0 # !!!
            else:
                punti_z_per_riga_3[i] = ((punti_t_per_riga_3[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)           
            punti_t_per_riga_3[i] = 50 + (10 * punti_z_per_riga_3[i])
            if punti_t_per_riga_3[i] < 40:
                punti_t_min_40_3.append(punti_t_per_riga_3[i])
            elif punti_t_per_riga_3[i] >= 40 and punti_t_per_riga_3[i] <= 60:
                punti_t_40_60_3.append(punti_t_per_riga_3[i])
            else:
                punti_t_magg_60_3.append(punti_t_per_riga_3[i])

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_fc)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_fc[i] = 0 # !!!
            else:
                punti_z_per_riga_fc[i] = ((punti_t_per_riga_fc[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga_fc[i] = 50 + (10 * punti_z_per_riga_fc[i])
            if punti_t_per_riga_fc[i] < 40:
                punti_t_min_40_fc.append(punti_t_per_riga_fc[i])
            elif punti_t_per_riga_fc[i] >= 40 and punti_t_per_riga_fc[i] <= 60:
                punti_t_40_60_fc.append(punti_t_per_riga_fc[i])
            else:
                punti_t_magg_60_fc.append(punti_t_per_riga_fc[i])

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_4)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_4[i] = 0
            else:
                punti_z_per_riga_4[i] = ((punti_t_per_riga_4[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga_4[i] = 50 + (10 * punti_z_per_riga_4[i])
            if punti_t_per_riga_4[i] < 40:
                punti_t_min_40_4.append(punti_t_per_riga_4[i])
            elif punti_t_per_riga_4[i] >= 40 and punti_t_per_riga_4[i] <= 60:
                punti_t_40_60_4.append(punti_t_per_riga_4[i])
            else:
                punti_t_magg_60_4.append(punti_t_per_riga_4[i])

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_5)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_5[i] = 0
            else:
                punti_z_per_riga_5[i] = ((punti_t_per_riga_5[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga_5[i] = 50 + (10 * punti_z_per_riga_5[i])
            if punti_t_per_riga_5[i] < 40:
                punti_t_min_40_5.append(punti_t_per_riga_5[i])
            elif punti_t_per_riga_5[i] >= 40 and punti_t_per_riga_5[i] <= 60:
                punti_t_40_60_5.append(punti_t_per_riga_5[i])
            else:
                punti_t_magg_60_5.append(punti_t_per_riga_5[i])

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_f2)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_f2[i] = 0
            else:
                punti_z_per_riga_f2[i] = ((punti_t_per_riga_f2[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga_f2[i] = 50 + (10 * punti_z_per_riga_f2[i])
            if punti_t_per_riga_f2[i] < 40:
                punti_t_min_40_f2.append(punti_t_per_riga_f2[i])
            elif punti_t_per_riga_f2[i] >= 40 and punti_t_per_riga_f2[i] <= 60:
                punti_t_40_60_f2.append(punti_t_per_riga_f2[i])
            else:
                punti_t_magg_60_f2.append(punti_t_per_riga_f2[i])

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_f3)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_f3[i] = 0
            else:
                punti_z_per_riga_f3[i] = ((punti_t_per_riga_f3[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga_f3[i] = 50 + (10 * punti_z_per_riga_f3[i])
            if punti_t_per_riga_f3[i] < 40:
                punti_t_min_40_f3.append(punti_t_per_riga_f3[i])
            elif punti_t_per_riga_f3[i] >= 40 and punti_t_per_riga_f3[i] <= 60:
                punti_t_40_60_f3.append(punti_t_per_riga_f3[i])
            else:
                punti_t_magg_60_f3.append(punti_t_per_riga_f3[i])
    
        # Calcola il punteggio medio dei punti T per la scala
        data[f'{dim}_min_40_2'] = punti_t_min_40_2
        data[f'{dim}_40_60_2'] = punti_t_40_60_2
        data[f'{dim}_magg_60_2'] = punti_t_magg_60_2
        data[f'{dim}_min_40_3'] = punti_t_min_40_3
        data[f'{dim}_40_60_3'] = punti_t_40_60_3
        data[f'{dim}_magg_60_3'] = punti_t_magg_60_3
        data[f'{dim}_min_40_fc'] = punti_t_min_40_fc
        data[f'{dim}_40_60_fc'] = punti_t_40_60_fc
        data[f'{dim}_magg_60_fc'] = punti_t_magg_60_fc
        data[f'{dim}_min_40_4'] = punti_t_min_40_4
        data[f'{dim}_40_60_4'] = punti_t_40_60_4
        data[f'{dim}_magg_60_4'] = punti_t_magg_60_4
        data[f'{dim}_min_40_5'] = punti_t_min_40_5
        data[f'{dim}_40_60_5'] = punti_t_40_60_5
        data[f'{dim}_magg_60_5'] = punti_t_magg_60_5
        data[f'{dim}_min_40_f2'] = punti_t_min_40_f2
        data[f'{dim}_40_60_f2'] = punti_t_40_60_f2
        data[f'{dim}_magg_60_f2'] = punti_t_magg_60_f2
        data[f'{dim}_min_40_f3'] = punti_t_min_40_f3
        data[f'{dim}_40_60_f3'] = punti_t_40_60_f3
        data[f'{dim}_magg_60_f3'] = punti_t_magg_60_f3

    return data

def get_exported_data_dips_1(file_csv, dip):
    df = pd.read_csv(file_csv, sep=';', dtype=object, encoding = 'ISO-8859-1')
    labs = {}

    file1 = openpyxl.load_workbook('files'+os.sep+'lab.xlsx', data_only=True)
    labs = file1['Laboratori e colloqui']

    n_org_tempo = 0
    n_strat_appr = 0
    n_gest_esame = 0
    n_colloqui = 0

    cdl_list = []
    for cdl, (dip_assoc, _) in cdl_dips_amb_dict.items():
        if dip_assoc == dip:
            cdl_list.append(cdls_dict[cdl])
            for row in labs.iter_rows(min_row=3, max_row=119, min_col=1, max_col=1, values_only=True):
                if row[0] == int(cdls_dict[cdl]):
                    n_org_tempo += labs.cell(row=row[0]+2, column=7).value if labs.cell(row=row[0]+2, column=7).value else 0
                    n_strat_appr += labs.cell(row=row[0]+2, column=8).value if labs.cell(row=row[0]+2, column=8).value else 0
                    n_gest_esame += labs.cell(row=row[0]+2, column=9).value if labs.cell(row=row[0]+2, column=9).value else 0
                    n_colloqui += labs.cell(row=row[0]+2, column=3).value if labs.cell(row=row[0]+2, column=3).value else 0

    if dip not in dips_dict:
        return f"Errore: dipartimento {dip} non trovato."
    
    cdl_list = []
    for cdl, (dip_assoc, _) in cdl_dips_amb_dict.items():
        if dip_assoc == dip:
            cdl_list.append(cdls_dict[cdl])
    df = df.loc[(df['Progress'] == '100')]
    df_1 = df[(df['CdS'].isin(cdl_list)) & (df['Anno'] == '1')]

    data = {
        'dip': dip,
        'dip_code': dips_dict[dip][1],
        'numero_studenti_1': df_1.shape[0],
        'numero_maschi_1': df_1.loc[df_1['GENERE'] == '1'].shape[0],
        'numero_femmine_1': df_1.loc[df_1['GENERE'] == '2'].shape[0],
        'numero_nonbin_1': df_1.loc[df_1['GENERE'] == '3'].shape[0],
        'numero_altro_1': df_1.loc[df_1['GENERE'] == '4'].shape[0],
        'numero_eta_media_1': round(pd.to_numeric(df_1['ETA'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_1': round(pd.to_numeric(df_1['ETA'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_1': df_1.loc[df_1['FREQ'] == '1'].shape[0],
        'no_partecip_1': df_1.loc[df_1['M1_NO'] == '1'].shape[0],

        'NUMLAB_1': n_org_tempo + n_strat_appr + n_gest_esame,
        'NUMCOLLOQUI_1': n_colloqui,

        'N_ORG_TEMPO_1': n_org_tempo,
        'N_STRAT_APPR_1': n_strat_appr,
        'N_GEST_ESAME_1': n_gest_esame

    }

    if np.isnan(data['numero_eta_media_1']):
        data['numero_eta_media_1'] = 0.0

    if np.isnan(data['deviazione_standard_1']):
        data['deviazione_standard_1'] = 0.0

    for dim in dimensions_dict_QPSS:
        # print("Dimensione:", dim)
        colonne_dimensione = dimensions_dict_QPSS[dim]
        minimo_teorico = 1 * len(colonne_dimensione)
        massimo_teorico = 4 * len(colonne_dimensione)

        punti_t_per_riga = []
        punti_t_min_40 = []
        punti_t_40_60 = []
        punti_t_magg_60 = []
        punti_z_per_riga = []
        punteggi_normalizzati_pop = []

        for index, row in df.iterrows():
            
            punteggi_grezzi_scala_pop = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                if pd.notna(row[colonna]):
                    if dim == "Gestire forme accentuate di ansietà":
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_pop.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_pop.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_pop.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_pop.append(1)
                    else:
                        punteggi_grezzi_scala_pop.append(int(row[colonna]))
                        
            punteggio_grezzo_totale_pop = sum(punteggi_grezzi_scala_pop)
            punteggio_normalizzato_pop = (punteggio_grezzo_totale_pop - minimo_teorico) / (massimo_teorico - minimo_teorico)
        
            punteggi_normalizzati_pop.append(punteggio_normalizzato_pop)
    
        # Itera attraverso le righe del dataframe
        for index, row in df_1.iterrows():
            
            punteggi_grezzi_scala = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if dim == "Gestire forme accentuate di ansietà":
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala.append(1)
                    else:
                        punteggi_grezzi_scala.append(int(row[colonna]))
                        
            punteggio_grezzo_totale = sum(punteggi_grezzi_scala)
            punteggio_normalizzato = (punteggio_grezzo_totale - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga.append(punteggio_normalizzato)  
            punti_t_per_riga.append(punteggio_normalizzato)
        
        media_punteggi_normalizzati_pop = np.mean(punteggi_normalizzati_pop)
        deviazione_standard_punteggi_normalizzati_pop = np.std(punteggi_normalizzati_pop)

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga[i] = 0 #!!!
            else:
                punti_z_per_riga[i] = ((punti_t_per_riga[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga[i] = 50 + 10 * punti_z_per_riga[i]
            if punti_t_per_riga[i] < 40:
                punti_t_min_40.append(punti_t_per_riga[i])
            elif punti_t_per_riga[i] >= 40 and punti_t_per_riga[i] <= 60:
                punti_t_40_60.append(punti_t_per_riga[i])
            else:
                punti_t_magg_60.append(punti_t_per_riga[i])
    
        data[f'{dim}_min_40_1'] = punti_t_min_40
        data[f'{dim}_40_60_1'] = punti_t_40_60
        data[f'{dim}_magg_60_1'] = punti_t_magg_60

    for dim in dimensions_dict_riflessioni:
        colonne_dimensione = dimensions_dict_riflessioni[dim]
        punteggi_per_nulla_daccordo = []
        punteggi_solo_in_parte_daccordo = []
        punteggi_abbastanza_daccordo = []
        punteggi_pienamente_daccordo = []

        for index, row in df_1.iterrows():
            for colonna in colonne_dimensione:
                if pd.notna(row[colonna]):
                    if (row[colonna] == "1"):
                        punteggi_per_nulla_daccordo.append(1)
                    elif (row[colonna] == "2"):
                        punteggi_solo_in_parte_daccordo.append(1)
                    elif (row[colonna] == "3"):
                        punteggi_abbastanza_daccordo.append(1)
                    elif (row[colonna] == "4"):
                        punteggi_pienamente_daccordo.append(1)
        
        data[f'{dim}_per_nulla_daccordo_1'] = len(punteggi_per_nulla_daccordo)
        data[f'{dim}_solo_in_parte_daccordo_1'] = len(punteggi_solo_in_parte_daccordo)
        data[f'{dim}_abbastanza_daccordo_1'] = len(punteggi_abbastanza_daccordo)
        data[f'{dim}_pienamente_daccordo_1'] = len(punteggi_pienamente_daccordo)

    return data

def get_exported_data_dips_23(file_csv, dip):
    df = pd.read_csv(file_csv, sep=';', dtype=object, encoding = 'ISO-8859-1')
    labs = {}

    file1 = openpyxl.load_workbook('files'+os.sep+'lab.xlsx', data_only=True)
    labs = file1['Laboratori e colloqui']

    n_org_tempo = 0
    n_gest_emo = 0
    n_pren_dec = 0
    n_facc_chiar = 0
    n_scelta_lm = 0
    n_personal_skills = 0
    n_colloqui_23 = 0
    
    # Funzione per pulire il valore della cella
    def clean_cell_value(value):
        if isinstance(value, str):
            value = value.replace('+', '').replace('\xa0', '').strip()
        return int(value) if value else 0

    cdl_list = []
    for cdl, (dip_assoc, _) in cdl_dips_amb_dict.items():
        if dip_assoc == dip:
            cdl_list.append(cdls_dict[cdl])
            # Itera sulle righe del foglio di lavoro
            for row in labs.iter_rows(min_row=3, max_row=119, min_col=1, max_col=1, values_only=True):
                if row[0] == int(cdls_dict[cdl]):
                    n_org_tempo += clean_cell_value(labs.cell(row=row[0]+2, column=10).value or 0) + \
                                clean_cell_value(labs.cell(row=row[0]+2, column=17).value or 0) + \
                                clean_cell_value(labs.cell(row=row[0]+2, column=18).value or 0)
        
                    n_gest_emo += (clean_cell_value(labs.cell(row=row[0]+2, column=11).value))
                    n_pren_dec += (clean_cell_value(labs.cell(row=row[0]+2, column=12).value or 0) + \
                                clean_cell_value(labs.cell(row=row[0]+2, column=16).value or 0))
                    n_facc_chiar += (clean_cell_value(labs.cell(row=row[0]+2, column=13).value))
                    n_scelta_lm += (clean_cell_value(labs.cell(row=row[0]+2, column=14).value))
                    n_personal_skills += (clean_cell_value(labs.cell(row=row[0]+2, column=15).value))
                    n_colloqui_23 += (clean_cell_value(labs.cell(row=row[0]+2, column=5).value))
    
    if dip not in dips_dict:
        return f"Errore: dipartimento {dip} non trovato."
    
    cdl_list = []
    for cdl, (dip_assoc, _) in cdl_dips_amb_dict.items():
        if dip_assoc == dip:
            cdl_list.append(cdls_dict[cdl])
    df = df.loc[(df['Progress'] == '100')]
    df_2 = df[(df['CdS'].isin(cdl_list)) & (df['ANNO'] == '2')]
    df_3 = df[(df['CdS'].isin(cdl_list)) & (df['ANNO'].isin(['3', '4', '5']))]
    df_fc = df[(df['CdS'].isin(cdl_list)) & (df['ANNO'] == '6')]
    df_3fc = df[(df['CdS'].isin(cdl_list)) & (df['ANNO'].isin(['3', '4', '5', '6']))]

    data = {
        'dip': dip,
        'dip_code': dips_dict[dip][1],
        'numero_studenti_2': df_2.shape[0],
        'numero_maschi_2': df_2.loc[df_2['GENERE'] == '1'].shape[0],
        'numero_femmine_2': df_2.loc[df_2['GENERE'] == '2'].shape[0],
        'numero_nonbin_2': df_2.loc[df_2['GENERE'] == '3'].shape[0],
        'numero_altro_2': df_2.loc[df_2['GENERE'] == '4'].shape[0],
        'numero_eta_media_2': round(pd.to_numeric(df_2['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_2': round(pd.to_numeric(df_2['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_2': df_2.loc[df_2['FREQ'] == '1'].shape[0],

        'numero_studenti_3': df_3.shape[0],
        'numero_maschi_3': df_3.loc[df_3['GENERE'] == '1'].shape[0],
        'numero_femmine_3': df_3.loc[df_3['GENERE'] == '2'].shape[0],
        'numero_nonbin_3': df_3.loc[df_3['GENERE'] == '3'].shape[0],
        'numero_altro_3': df_3.loc[df_3['GENERE'] == '4'].shape[0],
        'numero_eta_media_3': round(pd.to_numeric(df_3['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_3': round(pd.to_numeric(df_3['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_3': df_3.loc[df_3['FREQ'] == '1'].shape[0],

        'numero_studenti_fc': df_fc.shape[0],
        'numero_maschi_fc': df_fc.loc[df_fc['GENERE'] == '1'].shape[0],
        'numero_femmine_fc': df_fc.loc[df_fc['GENERE'] == '2'].shape[0],
        'numero_nonbin_fc': df_fc.loc[df_fc['GENERE'] == '3'].shape[0],
        'numero_altro_fc': df_fc.loc[df_fc['GENERE'] == '4'].shape[0],
        'numero_eta_media_fc': round(pd.to_numeric(df_fc['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_fc': round(pd.to_numeric(df_fc['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_fc': df_fc.loc[df_fc['FREQ'] == '1'].shape[0],

        'numero_eta_media_3fc': round(pd.to_numeric(df_3fc['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        
        'NUMLAB_23': n_org_tempo + n_gest_emo + n_pren_dec + n_facc_chiar + n_scelta_lm + n_personal_skills,
        'NUMCOLLOQUI_23': n_colloqui_23,

        'N_ORG_TEMPO_23': n_org_tempo,
        'N_GEST_EMO_23': n_gest_emo,
        'N_PREN_DEC_23': n_pren_dec,
        'N_FACC_CHIAR_23': n_facc_chiar,
        'N_SCELTA_LM_23': n_scelta_lm,
        'N_PERSONAL_SKILLS_23': n_personal_skills

    }

    if np.isnan(data['numero_eta_media_2']):
        data['numero_eta_media_2'] = 0.0

    if np.isnan(data['deviazione_standard_2']):
        data['deviazione_standard_2'] = 0.0
    
    if np.isnan(data['numero_eta_media_3']):
        data['numero_eta_media_2'] = 0.0

    if np.isnan(data['deviazione_standard_3']):
        data['deviazione_standard_2'] = 0.0

    if np.isnan(data['numero_eta_media_fc']):
        data['numero_eta_media_fc'] = 0.0

    if np.isnan(data['deviazione_standard_fc']):
        data['deviazione_standard_fc'] = 0.0

    for dim in dimensions_dict_QBEAP:
        # print("Dimensione:", dim)
        colonne_dimensione = dimensions_dict_QBEAP[dim]
        minimo_teorico = 1 * len(colonne_dimensione)
        massimo_teorico = 5 * len(colonne_dimensione)
        punteggi_normalizzati_pop = []

        # Punti T e Z secondo anno
        punti_t_per_riga_2 = []
        punti_t_min_40_2 = []
        punti_t_40_60_2 = []
        punti_t_magg_60_2 = []
        punti_z_per_riga_2 = []

        # Punti T e Z terzo anno
        punti_t_per_riga_3 = []
        punti_t_min_40_3 = []
        punti_t_40_60_3 = []
        punti_t_magg_60_3 = []
        punti_z_per_riga_3 = []

        # Punti T e Z fuori corso
        punti_t_per_riga_fc = []
        punti_t_min_40_fc = []
        punti_t_40_60_fc = []
        punti_t_magg_60_fc = []
        punti_z_per_riga_fc = []
    
        # Itera attraverso le righe del dataframe
        for index, row in df.iterrows():
            punteggi_grezzi_scala_pop = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_pop.append(5)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_pop.append(4)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_pop.append(3)    
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_pop.append(2)
                        elif (row[colonna] == "5"):
                            punteggi_grezzi_scala_pop.append(1)
                    else:
                        if row[colonna].strip().isdigit():
                            punteggi_grezzi_scala_pop.append(int(row[colonna]))
                        else:
                            punteggi_grezzi_scala_pop.append(0)
                        
            punteggio_grezzo_totale_pop = sum(punteggi_grezzi_scala_pop)
            punteggio_normalizzato_pop = (punteggio_grezzo_totale_pop - minimo_teorico) / (massimo_teorico - minimo_teorico)

            punteggi_normalizzati_pop.append(punteggio_normalizzato_pop)

        # Itera attraverso le righe del dataframe
        for index, row in df_2.iterrows():
            
            punteggi_grezzi_scala_2 = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_2.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_2.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_2.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_2.append(1)
                    else:
                        if row[colonna].strip().isdigit():
                            punteggi_grezzi_scala_2.append(int(row[colonna]))
                        else:
                            punteggi_grezzi_scala_2.append(0)
                        
            punteggio_grezzo_totale_2 = sum(punteggi_grezzi_scala_2)
            punteggio_normalizzato_2 = (punteggio_grezzo_totale_2 - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga_2.append(punteggio_normalizzato_2)  
            punti_t_per_riga_2.append(punteggio_normalizzato_2)

        # Itera attraverso le righe del dataframe
        for index, row in df_3.iterrows():
            
            punteggi_grezzi_scala_3 = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_3.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_3.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_3.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_3.append(1)
                    else:
                        if row[colonna].strip().isdigit():
                            punteggi_grezzi_scala_3.append(int(row[colonna]))
                        else:
                            punteggi_grezzi_scala_3.append(0)
                        
            punteggio_grezzo_totale_3 = sum(punteggi_grezzi_scala_3)
            punteggio_normalizzato_3 = (punteggio_grezzo_totale_3 - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga_3.append(punteggio_normalizzato_3)  
            punti_t_per_riga_3.append(punteggio_normalizzato_3)

        # Itera attraverso le righe del dataframe
        for index, row in df_fc.iterrows():
                
                punteggi_grezzi_scala_fc = []
    
                # Calcola il punteggio grezzo totale della scala
                for colonna in colonne_dimensione:
                    # print("Colonna:", colonna)
                    if pd.notna(row[colonna]):
                        if (colonna in ['GR_MIND_1', 'GR_MIND_4', 'GR_MIND_6', 'GR_MIND_8']
                        or colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_5', 'IIS_6', 'IIS_7', 'IIS_13', 'IIS_22']):
                            if (row[colonna] == "1"):
                                punteggi_grezzi_scala_fc.append(4)
                            elif (row[colonna] == "2"):
                                punteggi_grezzi_scala_fc.append(3)
                            elif (row[colonna] == "3"):
                                punteggi_grezzi_scala_fc.append(2)
                            elif (row[colonna] == "4"):
                                punteggi_grezzi_scala_fc.append(1)
                        else:
                            if row[colonna].strip().isdigit():
                                punteggi_grezzi_scala_fc.append(int(row[colonna]))
                            else:
                                punteggi_grezzi_scala_fc.append(0)
                            
                punteggio_grezzo_totale_fc = sum(punteggi_grezzi_scala_fc)
                punteggio_normalizzato_fc = (punteggio_grezzo_totale_fc - minimo_teorico) / (massimo_teorico - minimo_teorico)      
                punti_z_per_riga_fc.append(punteggio_normalizzato_fc)  
                punti_t_per_riga_fc.append(punteggio_normalizzato_fc)
        
        # Calcola il punteggio z utilizzando media e deviazione standard della popolazione (Ateneo)
        media_punteggi_normalizzati_pop = np.mean(punteggi_normalizzati_pop)
        deviazione_standard_punteggi_normalizzati_pop = np.std(punteggi_normalizzati_pop)

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_2)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_2[i] = 0 # !!!
            else:
                punti_z_per_riga_2[i] = ((punti_t_per_riga_2[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)           
            punti_t_per_riga_2[i] = 50 + 10 * punti_z_per_riga_2[i]
            if punti_t_per_riga_2[i] < 40:
                punti_t_min_40_2.append(punti_t_per_riga_2[i])
            elif punti_t_per_riga_2[i] >= 40 and punti_t_per_riga_2[i] <= 60:
                punti_t_40_60_2.append(punti_t_per_riga_2[i])
            else:
                punti_t_magg_60_2.append(punti_t_per_riga_2[i])          

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_3)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_3[i] = 0 # !!!
            else:
                punti_z_per_riga_3[i] = ((punti_t_per_riga_3[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)           
            punti_t_per_riga_3[i] = 50 + 10 * punti_z_per_riga_3[i]
            if punti_t_per_riga_3[i] < 40:
                punti_t_min_40_3.append(punti_t_per_riga_3[i])
            elif punti_t_per_riga_3[i] >= 40 and punti_t_per_riga_3[i] <= 60:
                punti_t_40_60_3.append(punti_t_per_riga_3[i])
            else:
                punti_t_magg_60_3.append(punti_t_per_riga_3[i])

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga_fc)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga_fc[i] = 0 # !!!
            else:
                punti_z_per_riga_fc[i] = ((punti_t_per_riga_fc[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga_fc[i] = 50 + 10 * punti_z_per_riga_fc[i]
            if punti_t_per_riga_fc[i] < 40:
                punti_t_min_40_fc.append(punti_t_per_riga_fc[i])
            elif punti_t_per_riga_fc[i] >= 40 and punti_t_per_riga_fc[i] <= 60:
                punti_t_40_60_fc.append(punti_t_per_riga_fc[i])
            else:
                punti_t_magg_60_fc.append(punti_t_per_riga_fc[i])
    
        # Calcola il punteggio medio dei punti T per la scala
        data[f'{dim}_min_40_2'] = punti_t_min_40_2
        data[f'{dim}_40_60_2'] = punti_t_40_60_2
        data[f'{dim}_magg_60_2'] = punti_t_magg_60_2
        data[f'{dim}_min_40_3'] = punti_t_min_40_3
        data[f'{dim}_40_60_3'] = punti_t_40_60_3
        data[f'{dim}_magg_60_3'] = punti_t_magg_60_3
        data[f'{dim}_min_40_fc'] = punti_t_min_40_fc
        data[f'{dim}_40_60_fc'] = punti_t_40_60_fc
        data[f'{dim}_magg_60_fc'] = punti_t_magg_60_fc

    return data

def get_exported_data_pot(file_csv, cdl_pot, uni_pot):
    df = pd.read_csv(file_csv, sep=';', dtype=object, encoding='ISO-8859-1')

    # CdS_1 è l'università di appartenenza
    # CdS_2 è il corso di laurea di appartenenza, cdl_pot
    
    if cdl_pot not in pot_cdl_dict:
        return f"Errore: corso di laurea {cdl_pot} non trovato"
    
    df = df.loc[(df['Progress'].isin(['90', '91', '92', '93', '94', '95', '96', '97', '98', '99', '100']))]
    
    if pot_cdl_dict[cdl_pot] not in cdl_composed_code_dict.values():
            for key, value in uni_cdl_dict.items():
                if cdl_pot in value:
                    uni_pot = key
            df_pot = df.loc[df['CdS_2'] == pot_cdl_dict[cdl_pot]]
    else:
        df_pot = df.loc[ (df['CdS_1'] == pot_uni_dict[uni_pot]) & (df['CdS_2'] == pot_cdl_dict[cdl_pot])]

    # print(df_pot)
    
    df_pot.loc[:, 'Esame_DIFF'] = df_pot['Esame_DIFF'].str.upper()
    exam_counts = df_pot['Esame_DIFF'].value_counts()
    top_3_exams = exam_counts.head(3)
    exams = top_3_exams.index.tolist()
    pop = df_pot.shape[0] if df_pot.shape[0] > 0 else 1

    data = {
        'cdl_pot': cdl_pot,
        'cdl_code_pot': pot_cdl_dict[cdl_pot],
        'uni_code_pot': uni_pot,
        'numero_studenti_pot': df_pot.shape[0],
        'numero_maschi_pot': df_pot.loc[df_pot['GENERE'] == '1'].shape[0],
        'numero_femmine_pot': df_pot.loc[df_pot['GENERE'] == '2'].shape[0],
        'numero_nonbin_pot': df_pot.loc[df_pot['GENERE'] == '3'].shape[0],
        'numero_altro_pot': df_pot.loc[df_pot['GENERE'] == '4'].shape[0],
        'numero_eta_media_pot': round(pd.to_numeric(df_pot['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_pot': round(pd.to_numeric(df_pot['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_pot': df_pot.loc[df_pot['FREQ'] == '1'].shape[0],
        'media_soddisf_pot': round(pd.to_numeric(df_pot['SODD_1'].str.replace('o','0'), errors='coerce').mean(), 2),
        'media_soddisf_pop': round(pd.to_numeric(df['SODD_1'].str.replace('o','0'), errors='coerce').mean(), 2),
        'percentuale_isolam_si_pot': round((df_pot.loc[df_pot['ISOLAM'] == '1'].shape[0] * 100)/pop, 2),

        'top_exam_1': exams[0] if len(exams) > 0 else None,
        'top_exam_2': exams[1] if len(exams) > 1 else None,
        'top_exam_3': exams[2] if len(exams) > 2 else None,
        'top_exam_1_percent': round((top_3_exams[exams[0]] * 100) / pop, 2) if len(exams) > 0 else None,
        'top_exam_2_percent': round((top_3_exams[exams[1]] * 100) / pop, 2) if len(exams) > 1 else None,
        'top_exam_3_percent': round((top_3_exams[exams[2]] * 100) / pop, 2) if len(exams) > 2 else None

    }

    if np.isnan(data['numero_eta_media_pot']):
        data['numero_eta_media_pot'] = 0

    if np.isnan(data['deviazione_standard_pot']):
        data['deviazione_standard_pot'] = 0

    for dim in dimensions_dict_POT:
        # print("Dimensione:", dim)
        colonne_dimensione = dimensions_dict_POT[dim]
        minimo_teorico = 1 * len(colonne_dimensione)
        massimo_teorico = 4 * len(colonne_dimensione)

        punti_t_per_riga = []
        punti_t_min_40 = []
        punti_t_40_60 = []
        punti_t_magg_60 = []
        punti_z_per_riga = []
        punteggi_normalizzati_pop = []

        for index, row in df.iterrows():
            
            punteggi_grezzi_scala_pop = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                if pd.notna(row[colonna]):
                    if (colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_1', 'IIS_2', 'IIS_3', 'IIS_9', 'IIS_13']
                        or colonna in ['HELICOPTER_5', 'HELICOPTER_14']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_pop.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_pop.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_pop.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_pop.append(1)
                    else:
                        colonna_value = row[colonna].strip()
                        if colonna_value and colonna_value.isdigit():
                            punteggi_grezzi_scala_pop.append(int(colonna_value))
                        
            punteggio_grezzo_totale_pop = sum(punteggi_grezzi_scala_pop)
            punteggio_normalizzato_pop = (punteggio_grezzo_totale_pop - minimo_teorico) / (massimo_teorico - minimo_teorico)
        
            punteggi_normalizzati_pop.append(punteggio_normalizzato_pop)
    
        # Itera attraverso le righe del dataframe
        for index, row in df_pot.iterrows():
            
            punteggi_grezzi_scala = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_1', 'IIS_2', 'IIS_3', 'IIS_9', 'IIS_13']
                        or colonna in ['HELICOPTER_5', 'HELICOPTER_14']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala.append(1)
                    else:
                        colonna_value = row[colonna].strip()
                        if colonna_value and colonna_value.isdigit():
                            punteggi_grezzi_scala.append(int(colonna_value))
                        
            punteggio_grezzo_totale = sum(punteggi_grezzi_scala)
            punteggio_normalizzato = (punteggio_grezzo_totale - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga.append(punteggio_normalizzato)  
            punti_t_per_riga.append(punteggio_normalizzato)
        
        media_punteggi_normalizzati_pop = np.mean(punteggi_normalizzati_pop)
        deviazione_standard_punteggi_normalizzati_pop = np.std(punteggi_normalizzati_pop)

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga[i] = 0 #!!!
            else:
                punti_z_per_riga[i] = ((punti_t_per_riga[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga[i] = 50 + (10 * punti_z_per_riga[i])
            if punti_t_per_riga[i] < 40:
                punti_t_min_40.append(punti_t_per_riga[i])
            elif punti_t_per_riga[i] >= 40 and punti_t_per_riga[i] <= 60:
                punti_t_40_60.append(punti_t_per_riga[i])
            else:
                punti_t_magg_60.append(punti_t_per_riga[i])
    
        data[f'{dim}_min_40_pot'] = punti_t_min_40
        data[f'{dim}_40_60_pot'] = punti_t_40_60
        data[f'{dim}_magg_60_pot'] = punti_t_magg_60

    return data

def get_exported_data_pot_uni(file_csv, uni_pot):
    df = pd.read_csv(file_csv, sep=';', dtype=object, encoding='ISO-8859-1')
    labs = {}

    # parte labs

    # CdS_1 è l'università di appartenenza
    # CdS_2 è il corso di laurea di appartenenza, cdl_pot
    
    df = df.loc[(df['Progress'].isin(['90', '91', '92', '93', '94', '95', '96', '97', '98', '99', '100']))]
    df_pot = df.loc[df['CdS_1'] == pot_uni_dict[uni_pot]]

    data = {
        'uni_pot': uni_pot,
        'uni_code_pot': pot_uni_dict[uni_pot],
        'uni_code_pot': uni_pot,
        'numero_studenti_pot': df_pot.shape[0],
        'numero_maschi_pot': df_pot.loc[df_pot['GENERE'] == '1'].shape[0],
        'numero_femmine_pot': df_pot.loc[df_pot['GENERE'] == '2'].shape[0],
        'numero_nonbin_pot': df_pot.loc[df_pot['GENERE'] == '3'].shape[0],
        'numero_altro_pot': df_pot.loc[df_pot['GENERE'] == '4'].shape[0],
        'numero_eta_media_pot': round(pd.to_numeric(df_pot['ET_'].str.replace('o','0'), errors='coerce').mean(), 2),
        'deviazione_standard_pot': round(pd.to_numeric(df_pot['ET_'].str.replace('o','0'), errors='coerce').std(), 2),
        'freq_pot': df_pot.loc[df_pot['FREQ'] == '1'].shape[0],
        'media_soddisf_pot': round(pd.to_numeric(df_pot['SODD_1'].str.replace('o','0'), errors='coerce').mean(), 2),
        'media_soddisf_pop': round(pd.to_numeric(df['SODD_1'].str.replace('o','0'), errors='coerce').mean(), 2),
        'percentuale_isolam_si_pot': round((df_pot.loc[df_pot['ISOLAM'] == '1'].shape[0] * 100)/df_pot.shape[0], 2),

    }

    if np.isnan(data['numero_eta_media_pot']):
        data['numero_eta_media_pot'] = 0

    if np.isnan(data['deviazione_standard_pot']):
        data['deviazione_standard_pot'] = 0

    if np.isnan(data['media_soddisf_pot']):
        data['media_soddisf_pot'] = 0

    if np.isnan(data['media_soddisf_pop']):
        data['media_soddisf_pop'] = 0

    if np.isnan(data['percentuale_isolam_si_pot']):
        data['percentuale_isolam_si_pot'] = 0

    for dim in dimensions_dict_POT:
        # print("Dimensione:", dim)
        colonne_dimensione = dimensions_dict_POT[dim]
        minimo_teorico = 1 * len(colonne_dimensione)
        massimo_teorico = 4 * len(colonne_dimensione)

        punti_t_per_riga = []
        punti_t_min_40 = []
        punti_t_40_60 = []
        punti_t_magg_60 = []
        punti_z_per_riga = []
        punteggi_normalizzati_pop = []

        for index, row in df.iterrows():
            
            punteggi_grezzi_scala_pop = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                if pd.notna(row[colonna]):
                    if (colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_1', 'IIS_2', 'IIS_3', 'IIS_9', 'IIS_13']
                        or colonna in ['HELICOPTER_5', 'HELICOPTER_14']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala_pop.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala_pop.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala_pop.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala_pop.append(1)
                    else:
                        colonna_value = row[colonna].strip()
                        if colonna_value and colonna_value.isdigit():
                            punteggi_grezzi_scala_pop.append(int(colonna_value))
                        
            punteggio_grezzo_totale_pop = sum(punteggi_grezzi_scala_pop)
            punteggio_normalizzato_pop = (punteggio_grezzo_totale_pop - minimo_teorico) / (massimo_teorico - minimo_teorico)
        
            punteggi_normalizzati_pop.append(punteggio_normalizzato_pop)
    
        # Itera attraverso le righe del dataframe
        for index, row in df_pot.iterrows():
            
            punteggi_grezzi_scala = []

            # Calcola il punteggio grezzo totale della scala
            for colonna in colonne_dimensione:
                # print("Colonna:", colonna)
                if pd.notna(row[colonna]):
                    if (colonna in ['RISP_SCAD_1', 'RISP_SCAD_2', 'RISP_SCAD_3', 'RISP_SCAD_4']
                        or colonna in ['IIS_1', 'IIS_2', 'IIS_3', 'IIS_9', 'IIS_13']
                        or colonna in ['HELICOPTER_5', 'HELICOPTER_14']):
                        if (row[colonna] == "1"):
                            punteggi_grezzi_scala.append(4)
                        elif (row[colonna] == "2"):
                            punteggi_grezzi_scala.append(3)
                        elif (row[colonna] == "3"):
                            punteggi_grezzi_scala.append(2)
                        elif (row[colonna] == "4"):
                            punteggi_grezzi_scala.append(1)
                    else:
                        colonna_value = row[colonna].strip()
                        if colonna_value and colonna_value.isdigit():
                            punteggi_grezzi_scala.append(int(colonna_value))
                        
            punteggio_grezzo_totale = sum(punteggi_grezzi_scala)
            punteggio_normalizzato = (punteggio_grezzo_totale - minimo_teorico) / (massimo_teorico - minimo_teorico)      
            punti_z_per_riga.append(punteggio_normalizzato)  
            punti_t_per_riga.append(punteggio_normalizzato)
        
        media_punteggi_normalizzati_pop = np.mean(punteggi_normalizzati_pop)
        deviazione_standard_punteggi_normalizzati_pop = np.std(punteggi_normalizzati_pop)

        # Itera attraverso i punteggi T per riga e trasforma in punteggi T
        for i in range(len(punti_t_per_riga)):
            if deviazione_standard_punteggi_normalizzati_pop == 0:
                punti_z_per_riga[i] = 0 #!!!
            else:
                punti_z_per_riga[i] = ((punti_t_per_riga[i] - media_punteggi_normalizzati_pop) 
                                / deviazione_standard_punteggi_normalizzati_pop)
            punti_t_per_riga[i] = 50 + (10 * punti_z_per_riga[i])
            if punti_t_per_riga[i] < 40:
                punti_t_min_40.append(punti_t_per_riga[i])
            elif punti_t_per_riga[i] >= 40 and punti_t_per_riga[i] <= 60:
                punti_t_40_60.append(punti_t_per_riga[i])
            else:
                punti_t_magg_60.append(punti_t_per_riga[i])
    
        data[f'{dim}_min_40_pot'] = punti_t_min_40
        data[f'{dim}_40_60_pot'] = punti_t_40_60
        data[f'{dim}_magg_60_pot'] = punti_t_magg_60

    return data

############################################################################################################
# Make plots
category_names = ['% < 40',
                  '% tra 40 e 60',
                  '% > 60']

category_rifl = ['Per nulla d\'accordo + Solo in parte d\'accordo', 'Abbastanza d\'accordo + Pienamente d\'accordo']

cdl_1_2_4 = [
    "Architettura (Cesena)",
    "Chimica e tecnologie farmaceutiche (Bologna)",
    "Conservazione e restauro dei beni culturali (Bologna)",
    "Farmacia (Bologna)",
    "Giurisprudenza (Bologna)",
    "Giurisprudenza (Ravenna)",
    "Medicina e chirurgia (Forlì)",
    "Medicina e chirurgia (Ravenna)",
    "Medicina veterinaria (Bologna)",
    "Pharmacy (Rimini)",
    "Scienze della formazione primaria (Bologna)"
]

cdl_1_3_5 = [
    "Medicina e chirurgia (Bologna)"
]

cdl_1_2_5 = [
    "Medicine and surgery (Bologna)"
]

def makeplot_1(pdf, data, x, y, dims_dict=dimensions_dict_QPSS, width=170):
    labels = []
    t_min_40_values = []
    t_40_60_values = []
    t_magg_60_values = []

    # Estrai i valori relativi ai punteggi per ogni dimensione
    for key in data.keys():
        if key.endswith('_min_40_1'):
            dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
            if dimension in dims_dict:  # Verifica se la dimensione è nel dizionario
                labels.append(dimension)
                t_min_40_values.append(data[key])
                t_40_60_values.append(data[f'{dimension}_40_60_1'])
                t_magg_60_values.append(data[f'{dimension}_magg_60_1'])

    # Inverti le liste per ottenere l'ordine desiderato sul grafico
    labels.reverse()
    t_min_40_values.reverse()
    t_40_60_values.reverse()
    t_magg_60_values.reverse()

    # Crea il grafico
    fig, ax = plt.subplots(figsize=(8, 11))

    # Colori delle categorie
    category_colors = ['#fe5d26', '#f2c078', '#faedca']

    # Inverti l'asse y, nascondi l'asse x e impostane i limiti
    ax.invert_yaxis()
    ax.xaxis.set_visible(False)
    ax.set_xticks(np.arange(0, 101, 10.0), labels=[])
    ax.xaxis.set_tick_params(bottom=False)

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)

    for i, label in enumerate(labels): 
        # Calcola le percentuali per le tre categorie
        total = len(t_min_40_values[i]) + len(t_40_60_values[i]) + len(t_magg_60_values[i])
        percentages = [
            (len(t_min_40_values[i]) * 100) / total,
            (len(t_40_60_values[i]) * 100) / total,
            (len(t_magg_60_values[i]) * 100) / total
        ]

        # Calcola le posizioni di partenza delle barre
        starts = [0, percentages[0], percentages[0] + percentages[1]]

        # Disegna le barre orizzontali
        for j in range(len(percentages)):
            ax.barh(label, percentages[j], left=starts[j], height=0.5, color=category_colors[j], edgecolor='black')

            # Aggiungi i numeri sopra le barre
            if percentages[j] != 0.0:
                ax.text(starts[j] + percentages[j] / 2, i, f"{percentages[j]:.1f}%", ha='center', va='center', color='black', fontweight='bold')
        

    # Crea le "fake" barre per ogni categoria con i colori corrispondenti
    fake_bar_min_40 = plt.bar([0], [0], color=category_colors[0], label=category_names[0])
    fake_bar_40_60 = plt.bar([0], [0], color=category_colors[1], label=category_names[1])
    fake_bar_magg_60 = plt.bar([0], [0], color=category_colors[2], label=category_names[2])

    # Aggiungi la legenda utilizzando le "fake" barre
    ax.legend(handles=[fake_bar_min_40, fake_bar_40_60, fake_bar_magg_60], ncol=3, bbox_to_anchor=(0, 1), loc='lower left', fontsize='small')

    # Salva il grafico come immagine
    fig.savefig('plot_1', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_1.png', w=width)
    os.remove('plot_1.png')
    plt.close('all')

def makeplot_23_page1(pdf, data, x, y, dims_dict=dimensions_dict_QBEAP, width=157):
    labels = []
    dimensions_page_1 = dims_dict.keys() - {'Interazione con docenti', 'Interesse dei docenti per lo sviluppo degli studenti e dell\'insegnamento',
                                            'Interazioni tra pari', 'Sviluppo accademico e intellettuale', 'Impegno verso gli obiettivi universitari',
                                            'Ripensamenti sul percorso universitario', 'Consapevolezza dei propri interessi professionali',
                                            'Chiarezza degli obiettivi professionali', 'Ricerca di informazioni sul futuro professionale'}

    if data['cdl'] in cdl_1_2_4:
        t_min_40_values = {'_2': [], '_4': [], '_f2': []}
        t_40_60_values = {'_2': [], '_4': [], '_f2': []}
        t_magg_60_values = {'_2': [], '_4': [], '_f2': []}
        for key in data.keys():
            if key.endswith(('_min_40_2', '_min_40_4', '_min_40_f2')):
                dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
                if dimension in dimensions_page_1:  # Verifica se la dimensione è nel dizionario
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_2', '_4', '_f2']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

        # Inverti le liste per ottenere l'ordine desiderato sul grafico
        labels.reverse()
        for suffix in ['_2', '_4', '_f2']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # Setta le dimensioni e la posizione delle barre
        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7  # Altezza di ogni barra
        index = np.arange(len(labels))

        # Definisci i colori delle categorie
        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_4 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_f2 = ['#66688f', '#e585ea', '#fecee9']

        # Disegna le barre divise in fasce
        bar_width = bar_height / 3  # Larghezza di ogni fascia

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_f2', '_4', '_2']):
                # Calcola le percentuali e le posizioni di inizio delle fasce
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                # Disegna le barre divise per fasce
                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_f2[k] if j == 0 else 
                            category_colors_4[k] if j == 1 else category_colors_2[k])
                    # Aggiungi le etichette sopra le barre
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)
                        
        # Aggiunge etichette, legende e personalizzazione dell'aspetto
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_f2 = [Patch(color=category_colors_f2[i]) for i in range(3)]
        legend_patches_4 = [Patch(color=category_colors_4[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

        # Aggiunge i titoli alle patch
        titles = ['Secondo Anno', 'Quarto Anno', 'Fuori Corso']
        start_x = 22
        for i, title in enumerate(titles):
            ax.text(start_x, 8.5, title, fontsize=8)
            start_x += 20

        # Combina le legende in un'unica legenda
        legend_patches =  legend_patches_2 + legend_patches_4 + legend_patches_f2
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    elif data['cdl'] in cdl_1_3_5:
        t_min_40_values = {'_3': [], '_5': [], '_f3': []}
        t_40_60_values = {'_3': [], '_5': [], '_f3': []}
        t_magg_60_values = {'_3': [], '_5': [], '_f3': []}
        for key in data.keys():
            if key.endswith(('_min_40_3', '_min_40_5', '_min_40_f3')):
                dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
                if dimension in dimensions_page_1:  # Verifica se la dimensione è nel dizionario
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_3', '_5', '_f3']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

        # Inverti le liste per ottenere l'ordine desiderato sul grafico
        labels.reverse()
        for suffix in ['_3', '_5', '_f3']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # Setta le dimensioni e la posizione delle barre
        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7  # Altezza di ogni barra
        index = np.arange(len(labels))

        # Definisci i colori delle categorie
        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_4 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_f2 = ['#66688f', '#e585ea', '#fecee9']

        # Disegna le barre divise in fasce
        bar_width = bar_height / 3  # Larghezza di ogni fascia

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_f3', '_5', '_3']):
                # Calcola le percentuali e le posizioni di inizio delle fasce
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                # Disegna le barre divise per fasce
                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_f2[k] if j == 0 else 
                            category_colors_4[k] if j == 1 else category_colors_2[k])
                    # Aggiungi le etichette sopra le barre
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)
                        
        # Aggiunge etichette, legende e personalizzazione dell'aspetto
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_f2 = [Patch(color=category_colors_f2[i]) for i in range(3)]
        legend_patches_4 = [Patch(color=category_colors_4[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

        # Aggiunge i titoli alle patch
        titles = ['Terzo Anno', 'Quinto Anno', 'Fuori Corso']
        start_x = 22
        for i, title in enumerate(titles):
            ax.text(start_x, 8.5, title, fontsize=8)
            start_x += 20

        # Combina le legende in un'unica legenda
        legend_patches =  legend_patches_2 + legend_patches_4 + legend_patches_f2
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    elif data['cdl'] in cdl_1_2_5:
        t_min_40_values = {'_2': [], '_5': [], '_f3': []}
        t_40_60_values = {'_2': [], '_5': [], '_f3': []}
        t_magg_60_values = {'_2': [], '_5': [], '_f3': []}
        for key in data.keys():
            if key.endswith(('_min_40_2', '_min_40_5', '_min_40_f3')):
                dimension = key[:-9]
                if dimension in dimensions_page_1:
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_2', '_5', '_f3']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])
        
        labels.reverse()
        for suffix in ['_2', '_5', '_f3']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # print(data[f'numero_studenti_2'], data[f'numero_studenti_5'], data[f'numero_studenti_f3'])

        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7
        index = np.arange(len(labels))

        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_5 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_f3 = ['#66688f', '#e585ea', '#fecee9']

        bar_width = bar_height / 3

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_f3', '_5', '_2']):
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_f3[k] if j == 0 else 
                            category_colors_5[k] if j == 1 else category_colors_2[k])
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)
                        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_f3 = [Patch(color=category_colors_f3[i]) for i in range(3)]
        legend_patches_5 = [Patch(color=category_colors_5[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]
        
        titles = ['Secondo Anno', 'Quinto Anno', 'Fuori Corso']
        start_x = 22
        for i, title in enumerate(titles):
            ax.text(start_x, 8.5, title, fontsize=8)
            start_x += 20
        
        legend_patches =  legend_patches_2 + legend_patches_5 + legend_patches_f3
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    elif data['cdl'] not in cdl_1_2_4 and data['cdl'] not in cdl_1_3_5 and data['cdl'] not in cdl_1_2_5:
        t_min_40_values = {'_2': [], '_3': [], '_fc': []}
        t_40_60_values = {'_2': [], '_3': [], '_fc': []}
        t_magg_60_values = {'_2': [], '_3': [], '_fc': []}
        for key in data.keys():
            if key.endswith(('_min_40_2', '_min_40_3', '_min_40_fc')):
                dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
                if dimension in dimensions_page_1:  # Verifica se la dimensione è nel dizionario
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_2', '_3', '_fc']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

        # Inverti le liste per ottenere l'ordine desiderato sul grafico
        labels.reverse()
        for suffix in ['_2', '_3', '_fc']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # Setta le dimensioni e la posizione delle barre
        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7  # Altezza di ogni barra
        index = np.arange(len(labels))

        # Definisci i colori delle categorie
        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_3 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_fc = ['#66688f', '#e585ea', '#fecee9']

        # Disegna le barre divise in fasce
        bar_width = bar_height / 3  # Larghezza di ogni fascia

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_fc', '_3', '_2']):
                # Calcola le percentuali e le posizioni di inizio delle fasce
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                # Disegna le barre divise per fasce
                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_fc[k] if j == 0 else 
                            category_colors_3[k] if j == 1 else category_colors_2[k])
                    # Aggiungi le etichette sopra le barre
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)

        # Aggiunge etichette, legende e personalizzazione dell'aspetto
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_fc = [Patch(color=category_colors_fc[i]) for i in range(3)]
        legend_patches_3 = [Patch(color=category_colors_3[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

        # Aggiunge i titoli alle patch
        titles = ['Secondo Anno', 'Terzo Anno', 'Fuori Corso']
        start_x = 22
        for i, title in enumerate(titles):
            ax.text(start_x, 8.5, title, fontsize=8)
            start_x += 20

        # Combina le legende in un'unica legenda
        legend_patches =  legend_patches_2 + legend_patches_3 + legend_patches_fc
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    # Salva il grafico come immagine
    fig1.savefig('plot_23', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_23.png', w=width)
    os.remove('plot_23.png')
    plt.close('all')

def makeplot_23_page2(pdf, data, x, y, dims_dict=dimensions_dict_QBEAP, width=170):
    labels = []
    dimensions_page_2 = dims_dict.keys() - {'Motivazione intrinseca', 'Motivazione identificata', 'Motivazione estrinseca',
                                            'Autoefficacia accademica', 'Mentalità di crescita', 'Consapevolezza del proprio apprendimento',
                                            'Rispetto delle scadenze', 'Benessere percepito'}

    if data['cdl'] in cdl_1_2_4:
        t_min_40_values = {'_2': [], '_4': [], '_f2': []}
        t_40_60_values = {'_2': [], '_4': [], '_f2': []}
        t_magg_60_values = {'_2': [], '_4': [], '_f2': []}
        for key in data.keys():
            if key.endswith(('_min_40_2', '_min_40_4', '_min_40_f2')):
                dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
                if dimension in dimensions_page_2:  # Verifica se la dimensione è nel dizionario
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_2', '_4', '_f2']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

        # Inverti le liste per ottenere l'ordine desiderato sul grafico
        labels.reverse()
        for suffix in ['_2', '_4', '_f2']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # Setta le dimensioni e la posizione delle barre
        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7  # Altezza di ogni barra
        index = np.arange(len(labels))

        # Definisci i colori delle categorie
        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_4 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_f2 = ['#66688f', '#e585ea', '#fecee9']

        # Disegna le barre divise in fasce
        bar_width = bar_height / 3  # Larghezza di ogni fascia

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_f2', '_4', '_2']):
                # Calcola le percentuali e le posizioni di inizio delle fasce
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                # Disegna le barre divise per fasce
                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_f2[k] if j == 0 else 
                            category_colors_4[k] if j == 1 else category_colors_2[k])
                    # Aggiungi le etichette sopra le barre
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)
                        
        # Aggiunge etichette, legende e personalizzazione dell'aspetto
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_f2 = [Patch(color=category_colors_f2[i]) for i in range(3)]
        legend_patches_4 = [Patch(color=category_colors_4[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

        # Aggiunge i titoli alle patch
        titles = ['Secondo Anno', 'Quarto Anno', 'Fuori Corso']
        start_x = 28
        for i, title in enumerate(titles):
            ax.text(start_x, 9.65, title, fontsize=8)
            start_x += 20

        # Combina le legende in un'unica legenda
        legend_patches =  legend_patches_2 + legend_patches_4 + legend_patches_f2
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    elif data['cdl'] in cdl_1_3_5:
        t_min_40_values = {'_3': [], '_5': [], '_f3': []}
        t_40_60_values = {'_3': [], '_5': [], '_f3': []}
        t_magg_60_values = {'_3': [], '_5': [], '_f3': []}
        for key in data.keys():
            if key.endswith(('_min_40_3', '_min_40_5', '_min_40_f3')):
                dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
                if dimension in dimensions_page_2:  # Verifica se la dimensione è nel dizionario
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_3', '_5', '_f3']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

        # Inverti le liste per ottenere l'ordine desiderato sul grafico
        labels.reverse()
        for suffix in ['_3', '_5', '_f3']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # Setta le dimensioni e la posizione delle barre
        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7  # Altezza di ogni barra
        index = np.arange(len(labels))

        # Definisci i colori delle categorie
        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_4 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_f2 = ['#66688f', '#e585ea', '#fecee9']

        # Disegna le barre divise in fasce
        bar_width = bar_height / 3  # Larghezza di ogni fascia

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_f3', '_5', '_3']):
                # Calcola le percentuali e le posizioni di inizio delle fasce
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                # Disegna le barre divise per fasce
                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_f2[k] if j == 0 else 
                            category_colors_4[k] if j == 1 else category_colors_2[k])
                    # Aggiungi le etichette sopra le barre
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)
                        
        # Aggiunge etichette, legende e personalizzazione dell'aspetto
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_f2 = [Patch(color=category_colors_f2[i]) for i in range(3)]
        legend_patches_4 = [Patch(color=category_colors_4[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

        # Aggiunge i titoli alle patch
        titles = ['Terzo Anno', 'Quinto Anno', 'Fuori Corso']
        start_x = 28
        for i, title in enumerate(titles):
            ax.text(start_x, 9.65, title, fontsize=8)
            start_x += 20

        # Combina le legende in un'unica legenda
        legend_patches =  legend_patches_2 + legend_patches_4 + legend_patches_f2
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.07), ncol=3, fontsize=8)

    elif data['cdl'] in cdl_1_2_5:
        t_min_40_values = {'_2': [], '_5': [], '_f3': []}
        t_40_60_values = {'_2': [], '_5': [], '_f3': []}
        t_magg_60_values = {'_2': [], '_5': [], '_f3': []}
        for key in data.keys():
            if key.endswith(('_min_40_2', '_min_40_5', '_min_40_f3')):
                dimension = key[:-9]
                if dimension in dimensions_page_2:
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_2', '_5', '_f3']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])
        
        labels.reverse()
        for suffix in ['_2', '_5', '_f3']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7
        index = np.arange(len(labels))

        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_5 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_f3 = ['#66688f', '#e585ea', '#fecee9']

        bar_width = bar_height / 3

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_f3', '_5', '_2']):
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_f3[k] if j == 0 else 
                            category_colors_5[k] if j == 1 else category_colors_2[k])
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)
                        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_f3 = [Patch(color=category_colors_f3[i]) for i in range(3)]
        legend_patches_5 = [Patch(color=category_colors_5[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]
        
        titles = ['Secondo Anno', 'Quinto Anno', 'Fuori Corso']
        start_x = 22
        for i, title in enumerate(titles):
            ax.text(start_x, 9.65, title, fontsize=8)
            start_x += 20
        
        legend_patches =  legend_patches_2 + legend_patches_5 + legend_patches_f3
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    elif data['cdl'] not in cdl_1_2_4 and data['cdl'] not in cdl_1_3_5 and data['cdl'] not in cdl_1_2_5:
        t_min_40_values = {'_2': [], '_3': [], '_fc': []}
        t_40_60_values = {'_2': [], '_3': [], '_fc': []}
        t_magg_60_values = {'_2': [], '_3': [], '_fc': []}
        for key in data.keys():
            if key.endswith(('_min_40_2', '_min_40_3', '_min_40_fc')):
                dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
                if dimension in dimensions_page_2:  # Verifica se la dimensione è nel dizionario
                    if dimension not in labels:
                        labels.append(dimension)
                        for suffix in ['_2', '_3', '_fc']:
                            t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                            t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                            t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

        # Inverti le liste per ottenere l'ordine desiderato sul grafico
        labels.reverse()
        for suffix in ['_2', '_3', '_fc']:
            t_min_40_values[suffix].reverse()
            t_40_60_values[suffix].reverse()
            t_magg_60_values[suffix].reverse()

        # Setta le dimensioni e la posizione delle barre
        fig1, ax = plt.subplots(figsize=(8, 11))
        bar_height = 0.7  # Altezza di ogni barra
        index = np.arange(len(labels))

        # Definisci i colori delle categorie
        category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
        category_colors_3 = ['#ea8f1f', '#f3d23f', '#f3e37c']
        category_colors_fc = ['#66688f', '#e585ea', '#fecee9']

        # Disegna le barre divise in fasce
        bar_width = bar_height / 3  # Larghezza di ogni fascia

        for i, label in enumerate(labels):
            for j, suffix in enumerate(['_fc', '_3', '_2']):
                # Calcola le percentuali e le posizioni di inizio delle fasce
                if data[f'numero_studenti{suffix}'] == 0:
                    total = 1
                else:
                    total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
                percentages = [
                    (len(t_min_40_values[suffix][i]) * 100) / total,
                    (len(t_40_60_values[suffix][i]) * 100) / total,
                    (len(t_magg_60_values[suffix][i]) * 100) / total
                ]
                starts = [0, percentages[0], percentages[0] + percentages[1]]

                # Disegna le barre divise per fasce
                for k in range(len(percentages)):
                    ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                            edgecolor='black', color=category_colors_fc[k] if j == 0 else 
                            category_colors_3[k] if j == 1 else category_colors_2[k])
                    # Aggiungi le etichette sopra le barre
                    if percentages[k] != 0.0:
                        ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                                ha='center', va='center', color='black', fontweight='bold', fontsize=8)

        # Aggiunge etichette, legende e personalizzazione dell'aspetto
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.set_yticks(index + 0.2)
        ax.xaxis.set_visible(False)
        ax.set_xticks([])
        ax.set_yticklabels(labels, fontsize=9, wrap=True)

        legend_patches_fc = [Patch(color=category_colors_fc[i]) for i in range(3)]
        legend_patches_3 = [Patch(color=category_colors_3[i]) for i in range(3)]
        legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

        # Aggiunge i titoli alle patch
        titles = ['Secondo Anno', 'Terzo Anno', 'Fuori Corso']
        start_x = 28
        for i, title in enumerate(titles):
            ax.text(start_x, 9.65, title, fontsize=8)
            start_x += 20

        # Combina le legende in un'unica legenda
        legend_patches =  legend_patches_2 + legend_patches_3 + legend_patches_fc
        legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

        ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)

    # Salva il grafico come immagine
    fig1.savefig('plot_23_2', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_23_2.png', w=width)
    os.remove('plot_23_2.png')
    plt.close('all')

def makeplot_dip23_page1(pdf, data, x, y, dims_dict=dimensions_dict_QBEAP, width=157):
    labels = []
    t_min_40_values = {'_2': [], '_3': [], '_fc': []}
    t_40_60_values = {'_2': [], '_3': [], '_fc': []}
    t_magg_60_values = {'_2': [], '_3': [], '_fc': []}
    dimensions_page_1 = dims_dict.keys() - {'Interazione con docenti', 'Interesse dei docenti per lo sviluppo degli studenti e dell\'insegnamento',
                                            'Interazioni tra pari', 'Sviluppo accademico e intellettuale', 'Impegno verso gli obiettivi universitari',
                                            'Ripensamenti sul percorso universitario', 'Consapevolezza dei propri interessi professionali',
                                            'Chiarezza degli obiettivi professionali', 'Ricerca di informazioni sul futuro professionale'}

    # print(data)

    for key in data.keys():
        if key.endswith(('_min_40_2', '_min_40_3', '_min_40_fc')):
            dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
            if dimension in dimensions_page_1:  # Verifica se la dimensione è nel dizionario
                if dimension not in labels:
                    labels.append(dimension)
                    for suffix in ['_2', '_3', '_fc']:
                        t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                        t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                        t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])

    # print(t_min_40_values)
    # print(t_40_60_values)
    # print(t_magg_60_values)

    # Inverti le liste per ottenere l'ordine desiderato sul grafico
    labels.reverse()
    for suffix in ['_2', '_3', '_fc']:
        t_min_40_values[suffix].reverse()
        t_40_60_values[suffix].reverse()
        t_magg_60_values[suffix].reverse()

    # Setta le dimensioni e la posizione delle barre
    fig1, ax = plt.subplots(figsize=(8, 11))
    bar_height = 0.7  # Altezza di ogni barra
    index = np.arange(len(labels))

    # Definisci i colori delle categorie
    category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
    category_colors_3 = ['#ea8f1f', '#f3d23f', '#f3e37c']
    category_colors_fc = ['#66688f', '#e585ea', '#fecee9']

    # Disegna le barre divise in fasce
    bar_width = bar_height / 3  # Larghezza di ogni fascia

    for i, label in enumerate(labels):
        for j, suffix in enumerate(['_fc', '_3', '_2']):
            # Calcola le percentuali e le posizioni di inizio delle fasce
            if data[f'numero_studenti{suffix}'] == 0:
                total = 1
            else:
                total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
            percentages = [
                (len(t_min_40_values[suffix][i]) * 100) / total,
                (len(t_40_60_values[suffix][i]) * 100) / total,
                (len(t_magg_60_values[suffix][i]) * 100) / total
            ]
            starts = [0, percentages[0], percentages[0] + percentages[1]]

            # Disegna le barre divise per fasce
            for k in range(len(percentages)):
                ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                        edgecolor='black', color=category_colors_fc[k] if j == 0 else 
                        category_colors_3[k] if j == 1 else category_colors_2[k])
                # Aggiungi le etichette sopra le barre
                if percentages[k] != 0.0:
                    ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                            ha='center', va='center', color='black', fontweight='bold', fontsize=8)

    # Aggiunge etichette, legende e personalizzazione dell'aspetto
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.set_yticks(index + 0.2)
    ax.xaxis.set_visible(False)
    ax.set_xticks([])
    ax.set_yticklabels(labels, fontsize=9, wrap=True)

    legend_patches_fc = [Patch(color=category_colors_fc[i]) for i in range(3)]
    legend_patches_3 = [Patch(color=category_colors_3[i]) for i in range(3)]
    legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

    # Aggiunge i titoli alle patch
    titles = ['Secondo Anno', 'Anni Successivi', 'Fuori Corso']
    start_x = 22
    for i, title in enumerate(titles):
        ax.text(start_x, 8.5, title, fontsize=8)
        start_x += 20

    # Combina le legende in un'unica legenda
    legend_patches =  legend_patches_2 + legend_patches_3 + legend_patches_fc
    legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

    ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.06), ncol=3, fontsize=8)
            
    # # Aggiusta i margini
    # plt.subplots_adjust(left=0.4, right=0.8, top=0.8, bottom=0.2)

    # # Mostra il grafico
    # plt.show()

    # Salva il grafico come immagine
    fig1.savefig('plot_23', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_23.png', w=width)
    os.remove('plot_23.png')
    plt.close('all')

def makeplot_dip23_page2(pdf, data, x, y, dims_dict=dimensions_dict_QBEAP, width=170):
    labels = []
    t_min_40_values = {'_2': [], '_3': [], '_fc': []}
    t_40_60_values = {'_2': [], '_3': [], '_fc': []}
    t_magg_60_values = {'_2': [], '_3': [], '_fc': []}
    dimensions_page_2 = dims_dict.keys() - {'Motivazione intrinseca', 'Motivazione identificata', 'Motivazione estrinseca',
                                            'Autoefficacia accademica', 'Mentalità di crescita', 'Consapevolezza del proprio apprendimento',
                                            'Rispetto delle scadenze', 'Benessere percepito'}

    for key in data.keys():
        if key.endswith(('_min_40_2', '_min_40_3', '_min_40_fc')):
            dimension = key[:-9]  # Rimuovi il suffisso per ottenere la dimensione
            if dimension in dimensions_page_2:  # Verifica se la dimensione è nel dizionario
                if dimension not in labels:
                    labels.append(dimension)
                    for suffix in ['_2', '_3', '_fc']:
                        t_min_40_values[suffix].append(data[f'{dimension}_min_40{suffix}'])
                        t_40_60_values[suffix].append(data[f'{dimension}_40_60{suffix}'])
                        t_magg_60_values[suffix].append(data[f'{dimension}_magg_60{suffix}'])
    # Inverti le liste per ottenere l'ordine desiderato sul grafico
    labels.reverse()
    for suffix in ['_2', '_3', '_fc']:
        t_min_40_values[suffix].reverse()
        t_40_60_values[suffix].reverse()
        t_magg_60_values[suffix].reverse()

    # Setta le dimensioni e la posizione delle barre
    fig2, ax = plt.subplots(figsize=(8, 11))
    bar_height = 0.7  # Altezza di ogni barra
    index = np.arange(len(labels))

    # Definisci i colori delle categorie
    category_colors_2 = ['#508a05', '#6cc551', '#9ffcdf']
    category_colors_3 = ['#ea8f1f', '#f3d23f', '#f3e37c']
    category_colors_fc = ['#66688f', '#e585ea', '#fecee9']

    # Disegna le barre divise in fasce
    bar_width = bar_height / 3  # Larghezza di ogni fascia

    for i, label in enumerate(labels):
        for j, suffix in enumerate(['_fc', '_3', '_2']):
            # Calcola le percentuali e le posizioni di inizio delle fasce
            if data[f'numero_studenti{suffix}'] == 0:
                total = 1
            else:
                total = len(t_min_40_values[suffix][i]) + len(t_40_60_values[suffix][i]) + len(t_magg_60_values[suffix][i])
            percentages = [
                (len(t_min_40_values[suffix][i]) * 100) / total,
                (len(t_40_60_values[suffix][i]) * 100) / total,
                (len(t_magg_60_values[suffix][i]) * 100) / total
            ]
            starts = [0, percentages[0], percentages[0] + percentages[1]]

            # Disegna le barre divise per fasce
            for k in range(len(percentages)):
                ax.barh(index[i] + j * bar_width, percentages[k], bar_width, left=starts[k],
                        edgecolor='black', color=category_colors_fc[k] if j == 0 else 
                        category_colors_3[k] if j == 1 else category_colors_2[k])
                # Aggiungi le etichette sopra le barre
                if percentages[k] != 0.0:
                    ax.text(starts[k] + percentages[k] / 2, index[i] + j * bar_width, f"{percentages[k]:.1f}%", 
                            ha='center', va='center', color='black', fontweight='bold', fontsize=8)

    # Aggiunge etichette, legende e personalizzazione dell'aspetto
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.set_yticks(index + 0.2)
    ax.xaxis.set_visible(False)
    ax.set_xticks([])
    ax.set_yticklabels(labels, fontsize=9, wrap=True)

    legend_patches_fc = [Patch(color=category_colors_fc[i]) for i in range(3)]
    legend_patches_3 = [Patch(color=category_colors_3[i]) for i in range(3)]
    legend_patches_2 = [Patch(color=category_colors_2[i]) for i in range(3)]

    # Aggiunge i titoli alle patch
    titles = ['Secondo Anno', 'Anni Successivi', 'Fuori Corso']
    start_x = 28
    for i, title in enumerate(titles):
        ax.text(start_x, 9.65, title, ha='center', va='bottom', fontsize=8)
        start_x += 20

    # Combina le legende in un'unica legenda
    legend_patches =  legend_patches_2 + legend_patches_3 + legend_patches_fc
    legend_labels =  [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)] + [f'{category_names[i]}' for i in range(3)]

    ax.legend(legend_patches, legend_labels, loc='upper center', bbox_to_anchor=(0.5, 1.07), ncol=3, fontsize=8)
            
    # # Aggiusta i margini
    # plt.subplots_adjust(left=0.4, right=0.8, top=0.8, bottom=0.2)

    # # Mostra il grafico
    # plt.show()

    # Salva il grafico come immagine
    fig2.savefig('plot_23_2', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_23_2.png', w=width)
    os.remove('plot_23_2.png')
    plt.close('all')

def makeplot_riflessioni(pdf, data, x, y, dims_dict=dimensions_dict_riflessioni, width=180):
    labels = []
    per_nulla_daccordo = []
    solo_in_parte_daccordo = []
    abbastanza_daccordo = []
    pienamente_daccordo = []

    for key in data.keys():
        if key.endswith('_per_nulla_daccordo_1'):
            dimension = key[:-21]
            if dimension in dimensions_dict_riflessioni:
                labels.append(dimension)
                per_nulla_daccordo.append(data[key])
                solo_in_parte_daccordo.append(data[f'{dimension}_solo_in_parte_daccordo_1'])
                abbastanza_daccordo.append(data[f'{dimension}_abbastanza_daccordo_1'])
                pienamente_daccordo.append(data[f'{dimension}_pienamente_daccordo_1'])

    fig, ax = plt.subplots(figsize=(10, 12))

    category_colors = ['#337ca0', '#fe5d26']

    ax.invert_yaxis()
    ax.xaxis.set_visible(False)
    ax.set_xticks(np.arange(0, 101, 10.0), labels=[])
    ax.xaxis.set_tick_params(bottom=False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.set_yticks(np.arange(len(labels)))
    ax.set_yticklabels(labels, fontsize=18)
    ax.set_title('Partecipare ai laboratori sulle strategie di studio è utile', fontsize=23, fontweight='bold', pad=35, loc='left', x=-0.5)

    for i, label in enumerate(labels):
        total = per_nulla_daccordo[i] + solo_in_parte_daccordo[i] + abbastanza_daccordo[i] + pienamente_daccordo[i]
        percentages = [
            ((per_nulla_daccordo[i] + solo_in_parte_daccordo[i]) / total) * 100,
            ((abbastanza_daccordo[i] + pienamente_daccordo[i]) / total) * 100
        ]
        starts = [0, percentages[0]]
        for j in range(len(percentages)):
            ax.barh(label, percentages[j], left=starts[j], height=0.3, color=category_colors[j], edgecolor='black')
            if percentages[j] != 0.0:
                ax.text(starts[j] + percentages[j] / 2, i, f"{percentages[j]:.1f}%", ha='center', va='center', color='black', fontweight='bold', fontsize=15)
    
    fake_bar_1 = plt.bar([0], [0], color=category_colors[0], label=category_rifl[0])
    fake_bar_2 = plt.bar([0], [0], color=category_colors[1], label=category_rifl[1])

    ax.legend(handles=[fake_bar_1, fake_bar_2], ncol=2, bbox_to_anchor=(0.5, -0.01), loc='upper center', fontsize=15)

    # Salva il grafico come immagine
    fig.savefig('plot_rif', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_rif.png', w=width)
    os.remove('plot_rif.png')
    plt.close('all')

def makeplot_23_pot(pdf, data, x, y, dims_dict=dimensions_dict_POT, width=175):
    labels = []
    t_min_40_values = []
    t_40_60_values = []
    t_magg_60_values = []

    for key in data.keys():
        if key.endswith('_min_40_pot'):
            dimension = key[:-11]  # Rimuovi il suffisso per ottenere la dimensione
            if dimension in dims_dict:  # Verifica se la dimensione è nel dizionario
                labels.append(dimension)
                t_min_40_values.append(data[key])
                t_40_60_values.append(data[f'{dimension}_40_60_pot'])
                t_magg_60_values.append(data[f'{dimension}_magg_60_pot'])

    # Crea il grafico
    fig, ax = plt.subplots(figsize=(8, 11))

    # Colori delle categorie
    category_colors = ['#fe5d26', '#f2c078', '#faedca']

    # Inverti l'asse y, nascondi l'asse x e impostane i limiti
    ax.invert_yaxis()
    ax.xaxis.set_visible(False)
    ax.set_xticks(np.arange(0, 101, 10.0), labels=[], wrap=True)
    ax.xaxis.set_tick_params(bottom=False)


    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)

    for i, label in enumerate(labels):
        # Calcola le percentuali per le tre categorie
        total = len(t_min_40_values[i]) + len(t_40_60_values[i]) + len(t_magg_60_values[i])
        percentages = [
            (len(t_min_40_values[i]) * 100) / total,
            (len(t_40_60_values[i]) * 100) / total,
            (len(t_magg_60_values[i]) * 100) / total
        ]

        # Calcola le posizioni di partenza delle barre
        starts = [0, percentages[0], percentages[0] + percentages[1]]

        # Disegna le barre orizzontali
        for j in range(len(percentages)):
            ax.barh(label, percentages[j], left=starts[j], height=0.5, color=category_colors[j], edgecolor='black')

            # Aggiungi i numeri sopra le barre
            if percentages[j] != 0.0:
                ax.text(starts[j] + percentages[j] / 2, i, f"{percentages[j]:.1f}%", ha='center', va='center', color='black', fontweight='bold')

    # Crea le "fake" barre per ogni categoria con i colori corrispondenti
    fake_bar_min_40 = plt.bar([0], [0], color=category_colors[0], label=category_names[0])
    fake_bar_40_60 = plt.bar([0], [0], color=category_colors[1], label=category_names[1])
    fake_bar_magg_60 = plt.bar([0], [0], color=category_colors[2], label=category_names[2])

    # Aggiungi la legenda utilizzando le "fake" barre
    ax.legend(handles=[fake_bar_min_40, fake_bar_40_60, fake_bar_magg_60], ncol=3, bbox_to_anchor=(0, 1), loc='lower left', fontsize='small')

    # plt.subplots_adjust(left=0.4, right=0.8, top=0.8, bottom=0.2)
    # plt.show()

    # Salva il grafico come immagine
    fig.savefig('plot_pot', bbox_inches='tight', transparent=True)
    pdf.set_xy(x, y)
    pdf.image('plot_pot.png', w=width)
    os.remove('plot_pot.png')
    plt.close('all')

def make_exam_table(pdf, data, x, y, width=177):
    # Creazione dei dati per la tabella
    exams = [data['top_exam_1'], data['top_exam_2'], data['top_exam_3']]
    percentages = [data['top_exam_1_percent'], data['top_exam_2_percent'], data['top_exam_3_percent']]

    # Creazione della rappresentazione grafica della tabella utilizzando Matplotlib
    fig, ax = plt.subplots(figsize=(8, 3))

    ax.axis('off')  # Nasconde gli assi

    # Aggiunta della tabella con colori delle celle e testo centrato
    table_data = [[exam, percent] for exam, percent in zip(exams, percentages)]

    # Aggiunta della tabella con colori delle celle e testo centrato
    table = ax.table(cellText=table_data, 
             colLabels=["Insegnamento", "%"], 
             loc='center', 
             cellLoc='center')
    
    cell_height = 0.15
    for i in range(len(table_data)):
        table._cells[(i, 0)]._text.set_fontsize(15)  
        table._cells[(i, 1)]._text.set_fontsize(15)
        table._cells[(i+1, 0)]._text.set_fontsize(13.5)
        table._cells[(i+1, 1)]._text.set_fontsize(13.5)
        table._cells[(i, 0)].set_height(cell_height)
        table._cells[(i, 1)].set_height(cell_height)
        table._cells[(i+1, 0)].set_height(cell_height)
        table._cells[(i+1, 1)].set_height(cell_height)
        if i == 0: 
                table._cells[(i, 0)].set_facecolor('lightblue')
                table._cells[(i, 0)]._text.set_weight('bold')
                table._cells[(i, 1)].set_facecolor('lightblue')
                table._cells[(i, 1)]._text.set_weight('bold')

    # Salva il grafico come immagine
    fig.savefig('table_pot.png', bbox_inches='tight', transparent=True)   
    pdf.set_xy(x, y)
    pdf.image('table_pot.png', w=width)
    os.remove('table_pot.png')
    plt.close('all')

############################################################################################################
# Make report
# Corsi di laurea
def report_1(pdf, data_1):
    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
    perc_freq_1 = round((data_1['freq_1'] * 100) / data_1['numero_studenti_1'], 2)
    txt = f"Hanno compilato il questionario {data_1['numero_studenti_1']} studenti (M = {data_1['numero_maschi_1']}; F = {data_1['numero_femmine_1']}; non binario = {data_1['numero_nonbin_1']}; Preferisco non specificarlo = {data_1['numero_altro_1']}) con età media di {data_1['numero_eta_media_1']:.2f} anni (DS = {data_1['deviazione_standard_1']:.2f}). Il {perc_freq_1:.2f}% risultano frequentanti.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dal CdS per il primo anno di corso, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
    pdf.multi_cell(180, 6, txt=txt)
    pdf.set_font('OSi', 'I', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 70) # Setta x e y in base alla pagina del report
    txt2 = " Si rimanda alla nota metodologica (pag.1) per l'interpretazione del grafico."
    pdf.multi_cell(180, 6, txt=txt2)

    makeplot_1(pdf, data_1, 15, 85)

    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 256)
    txt = f"A seguito della restituzione in aula, si è proceduto con l'attivazione delle proposte laboratoriali e di colloquio, descritte nella pagina successiva. I colloqui individuali attivati sono stati {data_1['NUMCOLLOQUI_1']}, le partecipazioni ai laboratori sono state {data_1['NUMLAB_1']}.\nI colloqui individuali dedicati alle strategie di studio erano disponibili anche per gli studenti di anni successivi al primo."
    pdf.multi_cell(180, 6, txt=txt)

def report_2_cdl(pdf, data_1):
    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(14.5, 49.5) # Setta x e y in base alla pagina del report
    perc_no_partecip = round((data_1['no_partecip_1'] * 100) / data_1['numero_studenti_1'], 2)
    txt = f'Ne risulta che il {perc_no_partecip}% NON ha mai partecipato ad attività di approfondimento sul metodo di studio.'
    pdf.multi_cell(180, 6, txt=txt)

    makeplot_riflessioni(pdf, data_1, 20, 90)

def report_23_page_1(pdf, data_23):
    
    if data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] != 0 and data_23['numero_studenti_4'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
        perc_freq_4 = round(((data_23['freq_4'] + data_23['freq_f2']) * 100) / (data_23['numero_studenti_4'] + data_23['numero_studenti_f2']), 2)
        txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAl 4° anno hanno compilato il questionario {data_23['numero_studenti_4'] + data_23['numero_studenti_f2']} studenti (M = {data_23['numero_maschi_4'] + data_23['numero_maschi_f2']}; F = {data_23['numero_femmine_4']+ data_23['numero_femmine_f2']}; non binario = {data_23['numero_nonbin_4'] + data_23['numero_nonbin_f2']}; Altro = {data_23['numero_altro_4'] + data_23['numero_altro_f2']}) con età media di {data_23['numero_eta_media_4f2']:.2f} anni (DS = {data_23['deviazione_standard_4']:.2f}). Il {perc_freq_4:.2f}% risulta frequentante. Hanno compilato {data_23['numero_studenti_f2']} persone fuori corso.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] == 0 and data_23['numero_studenti_4'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_4 = round(((data_23['freq_4'] + data_23['freq_f2']) * 100) / (data_23['numero_studenti_4'] + data_23['numero_studenti_f2']), 2)
        txt = f"Al 2° anno non è stato compilato il questionario.\n\nAl 4° anno hanno compilato il questionario {data_23['numero_studenti_4'] + data_23['numero_studenti_f2']} studenti (M = {data_23['numero_maschi_4'] + data_23['numero_maschi_f2']}; F = {data_23['numero_femmine_4']+ data_23['numero_femmine_f2']}; non binario = {data_23['numero_nonbin_4'] + data_23['numero_nonbin_f2']}; Altro = {data_23['numero_altro_4'] + data_23['numero_altro_f2']}) con età media di {data_23['numero_eta_media_4f2']:.2f} anni (DS = {data_23['deviazione_standard_4']:.2f}). Il {perc_freq_4:.2f}% risulta frequentante. Hanno compilato {data_23['numero_studenti_f2']} persone fuori corso.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] != 0 and data_23['numero_studenti_4'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
        txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAl 4° anno non è stato compilato il questionario.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] == 0 and data_23['numero_studenti_4'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        txt = f"Al 2° anno non è stato compilato il questionario.\n\nAl 4° anno non è stato compilato il questionario."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] != 0 and data_23['numero_studenti_5'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_3 = round((data_23['freq_3'] * 100) / data_23['numero_studenti_3'], 2)
        perc_freq_5 = round(((data_23['freq_5'] + data_23['freq_f3']) * 100) / (data_23['numero_studenti_5'] + data_23['numero_studenti_f3']), 2)
        txt = f"Per quanto riguarda il 3° anno, hanno compilato il questionario {data_23['numero_studenti_3']} studenti (M = {data_23['numero_maschi_3']}; F = {data_23['numero_femmine_3']}; non binario = {data_23['numero_nonbin_3']}; Altro = {data_23['numero_altro_3']}) con età media di {data_23['numero_eta_media_3']:.2f} anni (DS = {data_23['deviazione_standard_3']:.2f}). Il {perc_freq_3:.2f}% risulta frequentante.\n\nAl 5° anno hanno compilato il questionario {data_23['numero_studenti_5'] + data_23['numero_studenti_f3']} studenti di cui {data_23['numero_studenti_f3']} fuori corso (M = {data_23['numero_maschi_5'] + data_23['numero_maschi_f3']}; F = {data_23['numero_femmine_5'] + data_23['numero_femmine_f3']}; non binario = {data_23['numero_nonbin_5'] + data_23['numero_nonbin_f3']}; Altro = {data_23['numero_altro_5'] + data_23['numero_altro_f3']}) con età media di {data_23['numero_eta_media_5f3']:.2f} anni (DS = {data_23['deviazione_standard_5']:.2f}). Il {perc_freq_5:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] == 0 and data_23['numero_studenti_5'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_5 = round(((data_23['freq_5'] + data_23['freq_f3']) * 100) / (data_23['numero_studenti_5'] + data_23['numero_studenti_f3']), 2)
        txt = f"Al 3° anno non è stato compilato il questionario.\n\nAl 5° anno hanno compilato il questionario {data_23['numero_studenti_5'] + data_23['numero_studenti_f3']} studenti di cui {data_23['numero_studenti_f3']} fuori corso (M = {data_23['numero_maschi_5'] + data_23['numero_maschi_f3']}; F = {data_23['numero_femmine_5'] + data_23['numero_femmine_f3']}; non binario = {data_23['numero_nonbin_5'] + data_23['numero_nonbin_f3']}; Altro = {data_23['numero_altro_5'] + data_23['numero_altro_f3']}) con età media di {data_23['numero_eta_media_5f3']:.2f} anni (DS = {data_23['deviazione_standard_5']:.2f}). Il {perc_freq_5:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] != 0 and data_23['numero_studenti_5'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_3 = round((data_23['freq_3'] * 100) / data_23['numero_studenti_3'], 2)
        txt = f"Per quanto riguarda il 3° anno, hanno compilato il questionario {data_23['numero_studenti_3']} studenti (M = {data_23['numero_maschi_3']}; F = {data_23['numero_femmine_3']}; non binario = {data_23['numero_nonbin_3']}; Altro = {data_23['numero_altro_3']}) con età media di {data_23['numero_eta_media_3']:.2f} anni (DS = {data_23['deviazione_standard_3']:.2f}). Il {perc_freq_3:.2f}% risulta frequentante.\n\nAl 5° anno non è stato compilato il questionario.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] == 0 and data_23['numero_studenti_5'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        txt = f"Al 3° anno non è stato compilato il questionario.\n\nAl 5° anno non è stato compilato il questionario."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 and data_23['numero_studenti_5'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
        perc_freq_5 = round(((data_23['freq_5'] + data_23['freq_f3']) * 100) / (data_23['numero_studenti_5'] + data_23['numero_studenti_f3']), 2)
        txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAl 5° anno hanno compilato il questionario {data_23['numero_studenti_5'] + data_23['numero_studenti_f3']} studenti di cui {data_23['numero_studenti_f3']} fuori corso (M = {data_23['numero_maschi_5'] + data_23['numero_maschi_f3']}; F = {data_23['numero_femmine_5'] + data_23['numero_femmine_f3']}; non binario = {data_23['numero_nonbin_5'] + data_23['numero_nonbin_f3']}; Altro = {data_23['numero_altro_5'] + data_23['numero_altro_f3']}) con età media di {data_23['numero_eta_media_5f3']:.2f} anni (DS = {data_23['deviazione_standard_5']:.2f}). Il {perc_freq_5:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] == 0 and data_23['numero_studenti_5'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_5 = round(((data_23['freq_5'] + data_23['freq_f3']) * 100) / (data_23['numero_studenti_5'] + data_23['numero_studenti_f3']), 2)
        txt = f"Al 2° anno non è stato compilato il questionario.\n\nAl 5° anno hanno compilato il questionario {data_23['numero_studenti_5'] + data_23['numero_studenti_f3']} studenti di cui {data_23['numero_studenti_f3']} fuori corso (M = {data_23['numero_maschi_5'] + data_23['numero_maschi_f3']}; F = {data_23['numero_femmine_5'] + data_23['numero_femmine_f3']}; non binario = {data_23['numero_nonbin_5'] + data_23['numero_nonbin_f3']}; Altro = {data_23['numero_altro_5'] + data_23['numero_altro_f3']}) con età media di {data_23['numero_eta_media_5f3']:.2f} anni (DS = {data_23['deviazione_standard_5']:.2f}). Il {perc_freq_5:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 and data_23['numero_studenti_5'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
        txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAl 5° anno non è stato compilato il questionario.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] == 0 and data_23['numero_studenti_5'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        txt = f"Al 2° anno non è stato compilato il questionario.\n\nAl 5° anno non è stato compilato il questionario."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['cdl'] not in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 and data_23['numero_studenti_3'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
        perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
        perc_freq_3 = round(((data_23['freq_3'] + data_23['freq_fc']) * 100) / (data_23['numero_studenti_3'] + data_23['numero_studenti_fc']), 2)
        txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAl 3° anno hanno compilato il questionario {data_23['numero_studenti_3'] + data_23['numero_studenti_fc']} studenti di cui {data_23['numero_studenti_fc']} fuori corso (M = {data_23['numero_maschi_3'] + data_23['numero_maschi_fc']}; F = {data_23['numero_femmine_3'] + data_23['numero_femmine_fc']}; non binario = {data_23['numero_nonbin_3'] + data_23['numero_nonbin_fc']}; Altro = {data_23['numero_altro_3'] + data_23['numero_altro_fc']}) con età media di {data_23['numero_eta_media_3fc']:.2f} anni (DS = {data_23['deviazione_standard_3']:.2f}). Il {perc_freq_3:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['cdl'] not in cdl_1_2_5 and data_23['numero_studenti_2'] == 0 and data_23['numero_studenti_3'] != 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
        perc_freq_3 = round(((data_23['freq_3'] + data_23['freq_fc']) * 100) / (data_23['numero_studenti_3'] + data_23['numero_studenti_fc']), 2)
        txt = f"Al 2° anno non è stato compilato il questionario.\n\nAl 3° anno hanno compilato il questionario {data_23['numero_studenti_3'] + data_23['numero_studenti_fc']} studenti di cui {data_23['numero_studenti_fc']} fuori corso (M = {data_23['numero_maschi_3'] + data_23['numero_maschi_fc']}; F = {data_23['numero_femmine_3'] + data_23['numero_femmine_fc']}; non binario = {data_23['numero_nonbin_3'] + data_23['numero_nonbin_3']}; Altro = {data_23['numero_altro_3'] + data_23['numero_altro_fc']}) con età media di {data_23['numero_eta_media_3fc']:.2f} anni (DS = {data_23['deviazione_standard_3']:.2f}). Il {perc_freq_3:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['cdl'] not in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 and data_23['numero_studenti_3'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
        perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
        txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAl 3° anno non è stato compilato il questionario.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
    elif data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['cdl'] not in cdl_1_2_5 and data_23['numero_studenti_2'] == 0 and data_23['numero_studenti_3'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        txt = f"Al 2° anno non è stato compilato il questionario.\n\nAl 3° anno non è stato compilato il questionario."
        pdf.multi_cell(180, 6, txt=txt)

    if (data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_4'] != 0) or \
    (data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] != 0 or data_23['numero_studenti_5'] != 0) or \
    (data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_5'] != 0) or \
    (data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_3'] != 0):
       
        pdf.set_font('OSi', 'I', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 85) # Setta x e y in base alla pagina del report
        txt2 = " Si rimanda alla nota metodologica (pag.1) per l'interpretazione del grafico."
        pdf.multi_cell(180, 6, txt=txt2)
        makeplot_23_page1(pdf, data_23, 23, 94) #regolare x e y

def report_23_page_2(pdf, data_23):
    if (data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_4'] != 0) or \
    (data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] != 0 or data_23['numero_studenti_5'] != 0) or \
    (data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_5'] != 0) or \
    (data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['cdl'] not in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_3'] != 0):
        makeplot_23_page2(pdf, data_23, 23, 50) #regolare x e y

        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 259)
        txt = f"A seguito della restituzione in aula, si è proceduto con l'attivazione delle proposte laboratoriali e di colloquio, descritte nella pagina successiva. I colloqui individuali attivati sono stati {data_23['NUMCOLLOQUI_23']} e hanno partecipato ai laboratori {data_23['NUMLAB_23']} persone."
        pdf.multi_cell(180, 6, txt=txt)

# Dipartimenti
def report_1_dip(pdf, data_1):
    pdf.set_font('OSr', 'R', 10.5)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(14.5, 32) # Setta x e y in base alla pagina del report
    perc_freq_1 = round((data_1['freq_1'] * 100) / data_1['numero_studenti_1'], 2)
    txt = f"Hanno compilato il questionario {data_1['numero_studenti_1']} persone (M = {data_1['numero_maschi_1']}; F = {data_1['numero_femmine_1']}; non binario = {data_1['numero_nonbin_1']}; Preferisco non specificarlo = {data_1['numero_altro_1']}) con età media di {data_1['numero_eta_media_1']} anni (DS = {data_1['deviazione_standard_1']}). Il {perc_freq_1:.1f}% risultano frequentanti.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dal CdS per il primo anno di corso, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
    pdf.multi_cell(180, 6, txt=txt)
    pdf.set_font('OSi', 'I', 10.5)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(14.5, 68) # Setta x e y in base alla pagina del report
    txt2 = "Si rimanda alla nota metodologica (pag.1) per l'interpretazione del grafico."
    pdf.multi_cell(180, 6, txt=txt2)

    makeplot_1(pdf, data_1, 15, 84)

def report_2_dip(pdf, data_1, data_23, dip):
    pdf.set_font('OSr', 'R', 10.5)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(14.5, 76.5) # Setta x e y in base alla pagina del report
    cdl_list = []
    txt = f"I dati resituiti in questo report fanno riferimento al Dipartimento di {clean(dip)}, di cui hanno partecipato al progetto i seguenti Corsi di Studio:\n"
    for cdl, (dip_assoc, _) in cdl_dips_amb_dict.items():
        if dip_assoc == dip:
            cdl_list.append(cdls_dict[cdl])
            txt += f" • {cdl}\t\t"
    pdf.multi_cell(175, 6, txt=txt)

    pdf.set_font('OSb', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(146, 175) # Setta x e y in base alla pagina del report
    txt2 = f'{data_1['numero_studenti_1']}'
    pdf.multi_cell(180, 6, txt=txt2)
    pdf.set_font('OSb', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(147, 192) # Setta x e y in base alla pagina del report
    txt2 = f'{data_1['NUMLAB_1']}'
    pdf.multi_cell(180, 6, txt=txt2)
    pdf.set_font('OSb', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(149, 209) # Setta x e y in base alla pagina del report
    txt2 = f'{data_1["NUMCOLLOQUI_1"]}'
    pdf.multi_cell(180, 6, txt=txt2)

    pdf.set_font('OSb', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(147, 241) # Setta x e y in base alla pagina del report
    numero_studenti_23 = data_23['numero_studenti_2'] + data_23['numero_studenti_3'] + data_23['numero_studenti_fc']
    txt3 = f'{numero_studenti_23}'
    pdf.multi_cell(180, 6, txt=txt3)
    pdf.set_font('OSb', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(149, 257.5) # Setta x e y in base alla pagina del report
    txt3 = f'{data_23['NUMLAB_23']}'
    pdf.multi_cell(180, 6, txt=txt3)
    pdf.set_font('OSb', 'B', 13)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(149, 274) # Setta x e y in base alla pagina del report
    txt3 = f'{data_23["NUMCOLLOQUI_23"]}'
    pdf.multi_cell(180, 6, txt=txt3)

def report_3_dip(pdf, data_1):
    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(177, 90)
    txt = f'{data_1['N_ORG_TEMPO_1']}'
    pdf.multi_cell(180, 6, txt=txt)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(177, 124)
    txt2 = f'{data_1['N_STRAT_APPR_1']}'
    pdf.multi_cell(180, 6, txt=txt2)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(177, 157.5)
    txt3 = f'{data_1['N_GEST_ESAME_1']}'
    pdf.multi_cell(180, 6, txt=txt3)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(177, 187)
    txt4 = f'{data_1['NUMCOLLOQUI_1']}'
    pdf.multi_cell(180, 6, txt=txt4)

def report_4_dip(pdf, data_1):
    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(14.5, 49.5) # Setta x e y in base alla pagina del report
    perc_no_partecip = round((data_1['no_partecip_1'] * 100) / data_1['numero_studenti_1'], 2)
    txt = f'Ne risulta che il {perc_no_partecip:.1f}% NON ha mai partecipato ad attività di approfondimento sul metodo di studio.'
    pdf.multi_cell(180, 6, txt=txt)

    makeplot_riflessioni(pdf, data_1, 20, 90)

def report_5_dip(pdf, data_23):
    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 81.5)
    txt = f'{data_23['N_ORG_TEMPO_23']}'
    pdf.multi_cell(180, 6, txt=txt)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 106)
    txt2 = f'{data_23['N_GEST_EMO_23']}'
    pdf.multi_cell(180, 6, txt=txt2)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 130.5)
    txt3 = f'{data_23['N_PREN_DEC_23']}'
    pdf.multi_cell(180, 6, txt=txt3)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 157)
    txt4 = f'{data_23['N_FACC_CHIAR_23']}'
    pdf.multi_cell(180, 6, txt=txt4)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 184.5)
    txt5 = f'{data_23['N_SCELTA_LM_23']}'
    pdf.multi_cell(180, 6, txt=txt5)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 207.5)
    txt6 = f'{data_23['N_PERSONAL_SKILLS_23']}'
    pdf.multi_cell(180, 6, txt=txt6)

    pdf.set_font('OSb', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(170, 226.5)
    txt7 = f'{data_23['NUMCOLLOQUI_23']}'
    pdf.multi_cell(180, 6, txt=txt7)

def report_23_page_1_dip(pdf, data_23):
    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
    perc_freq_2 = round((data_23['freq_2'] * 100) / data_23['numero_studenti_2'], 2)
    perc_freq_3 = round(((data_23['freq_3']+data_23['freq_fc']) * 100) / (data_23['numero_studenti_3']+data_23['numero_studenti_fc']), 2)
    txt = f"Per quanto riguarda il 2° anno, hanno compilato il questionario {data_23['numero_studenti_2']} studenti (M = {data_23['numero_maschi_2']}; F = {data_23['numero_femmine_2']}; non binario = {data_23['numero_nonbin_2']}; Altro = {data_23['numero_altro_2']}) con età media di {data_23['numero_eta_media_2']:.2f} anni (DS = {data_23['deviazione_standard_2']:.2f}). Il {perc_freq_2:.2f}% risulta frequentante.\n\nAgli anni successivi al 2° hanno compilato il questionario {data_23['numero_studenti_3'] + data_23['numero_studenti_fc']} studenti, di cui {data_23['numero_studenti_fc']} fuori corso (M = {data_23['numero_maschi_3'] + data_23['numero_maschi_fc']}; F = {data_23['numero_femmine_3']+ data_23['numero_femmine_fc']}; non binario = {data_23['numero_nonbin_3']+data_23['numero_nonbin_fc']}; Altro = {data_23['numero_altro_3']+data_23['numero_altro_fc']}) con età media di {data_23['numero_eta_media_3fc']:.2f} anni (DS = {data_23['deviazione_standard_3']:.2f}). Il {perc_freq_3:.2f}% risulta frequentante.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
    pdf.multi_cell(180, 6, txt=txt)
    pdf.set_font('OSi', 'I', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 85.5) # Setta x e y in base alla pagina del report
    txt2 = " Si rimanda alla nota metodologica (pag.1) per l'interpretazione del grafico."
    pdf.multi_cell(180, 6, txt=txt2)

    makeplot_dip23_page1(pdf, data_23, 23, 94) #regolare x e y

def report_23_page_2_dip(pdf, data_23): 
    makeplot_dip23_page2(pdf, data_23, 23, 50) #regolare x e y

# Corsi di laurea POT
def report_pot(pdf, data_pot):
    if data_pot['numero_studenti_pot'] == 0:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32)
        txt = f"Non è stato compilato il questionario."
        pdf.multi_cell(180, 6, txt=txt)
    else:
        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
        perc_freq_pot = round((data_pot['freq_pot'] * 100) / data_pot['numero_studenti_pot'], 2)
        txt = f"Il questionario è stato compilato da {data_pot['numero_studenti_pot']} studenti (M = {data_pot['numero_maschi_pot']}; F = {data_pot['numero_femmine_pot']}; non binario = {data_pot['numero_nonbin_pot']}; Altro = {data_pot['numero_altro_pot']}) iscritti al primo anno, con età media di {data_pot['numero_eta_media_pot']} anni (DS = {data_pot['deviazione_standard_pot']}). Il {perc_freq_pot}% risultano frequentanti.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
        pdf.multi_cell(180, 6, txt=txt)
        pdf.set_font('OSi', 'I', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(14.5, 62) # Setta x e y in base alla pagina del report
        txt2 = " Si rimanda alla nota metodologica (pag.1) per l'interpretazione del grafico."
        pdf.multi_cell(180, 6, txt=txt2)

        makeplot_23_pot(pdf, data_pot, 15, 85)

        pdf.set_font('OSr', 'R', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(15, 252)
        txt3 = f"È stato chiesto agli studenti di indicare quanto si sentissero soddisfatti del percorso intrapreso su una scala da 1 a 10. La media dei punteggi ottenuti dai partecipanti del corso è pari a {data_pot['media_soddisf_pot']}, mentre quella dei punteggi emersi da tutte le compilazioni è {data_pot['media_soddisf_pop']}.\n\nInoltre, con il fine di rilevare la presenza di una rete sociale per gli studenti, è stato chiesto loro di indicare se si sentissero socialmente isolati: {data_pot['percentuale_isolam_si_pot']:.1f}% ha risposto in modo affermativo."
        pdf.multi_cell(0, 6, txt=txt3)

def report_table(pdf, data_pot):
    make_exam_table(pdf, data_pot, 20, 70)

# Università POT
def report_pot_uni(pdf, data_pot):
    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 32) # Setta x e y in base alla pagina del report
    perc_freq_pot = round((data_pot['freq_pot'] * 100) / data_pot['numero_studenti_pot'], 2)
    txt = f"Il questionario è stato compilato da {data_pot['numero_studenti_pot']} studenti (M = {data_pot['numero_maschi_pot']}; F = {data_pot['numero_femmine_pot']}; non binario = {data_pot['numero_nonbin_pot']}; Altro = {data_pot['numero_altro_pot']}) iscritti al primo anno, con età media di {data_pot['numero_eta_media_pot']} anni (DS = {data_pot['deviazione_standard_pot']}). Il {perc_freq_pot}% risultano frequentanti.\n\nDi seguito sono riportati, per ciascuna dimensione, i valori relativi alla media dei punteggi ottenuti dai rispondenti, confrontati con la media dei punteggi di Ateneo (fissata convenzionalmente a 50)."
    pdf.multi_cell(180, 6, txt=txt)
    pdf.set_font('OSi', 'I', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(14.5, 62) # Setta x e y in base alla pagina del report
    txt2 = " Si rimanda alla nota metodologica (pag.1) per l'interpretazione del grafico."
    pdf.multi_cell(180, 6, txt=txt2)

    makeplot_23_pot(pdf, data_pot, 15, 85)

    pdf.set_font('OSr', 'R', 11)
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(15, 252)
    txt3 = f"È stato chiesto agli studenti di indicare quanto si sentissero soddisfatti del percorso intrapreso su una scala da 1 a 10. La media dei punteggi ottenuti dai partecipanti del corso è pari a {data_pot['media_soddisf_pot']}, mentre quella dei punteggi emersi da tutte le compilazioni è {data_pot['media_soddisf_pop']}.\n\nInoltre, con il fine di rilevare la presenza di una rete sociale per gli studenti, è stato chiesto loro di indicare se si sentissero socialmente isolati: {data_pot['percentuale_isolam_si_pot']:.1f}% ha risposto in modo affermativo."
    pdf.multi_cell(0, 6, txt=txt3)

############################################################################################################
# Create pdf
def create_pdf_cdls(file_1, file_23, cdl):
    data_1 = get_exported_data_cdls_1(file_1, cdl)
    data_23 = get_exported_data_cdls_23(file_23, cdl)
    
    pdf = PDF()
    pdf.add_fonts()

    # P1
    pdf.add_page()
    pdf.bg('p1', 'png_cdls')
    pdf.set_xy(10,195)     
    pdf.set_font('OSb', 'B', 23)
    pdf.set_text_color(164, 36, 70)

    pdf.multi_cell(0, 12, align='L', txt=f"REPORT CONCLUSIVO DEL CORSO\nDI STUDIO in {clean(cdl)}")

    # P2
    pdf.add_page()
    pdf.bg('p2', 'png_cdls')

    # P3
    pdf.add_page()
    pdf.bg('p3', 'png_cdls')

    # P4
    pdf.add_page()
    pdf.bg('p4', 'png_cdls')

    # P5 
    pdf.add_page()
    pdf.bg('p5', 'png_cdls')
    
    # P6
    pdf.add_page()
    pdf.bg('p6', 'png_cdls')

    # P7
    pdf.add_page()
    pdf.bg('p7', 'png_cdls')
    report_1(pdf, data_1)

    # P8
    pdf.add_page()
    pdf.bg('p8', 'png_cdls')

    # P9
    pdf.add_page()
    pdf.bg('p9', 'png_cdls')
    report_2_cdl(pdf, data_1)

    # P10
    pdf.add_page()
    pdf.bg('p10', 'png_cdls')

    # P11
    pdf.add_page()
    pdf.bg('p11', 'png_cdls')

    # P12
    pdf.add_page()
    pdf.bg('p12', 'png_cdls')
    report_23_page_1(pdf, data_23)

    # P13
    if (data_23['cdl'] in cdl_1_2_4 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_4'] != 0) or \
    (data_23['cdl'] in cdl_1_3_5 and data_23['numero_studenti_3'] != 0 or data_23['numero_studenti_5'] != 0) or \
    (data_23['cdl'] in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_5'] != 0) or \
    (data_23['cdl'] not in cdl_1_2_4 and data_23['cdl'] not in cdl_1_3_5 and data_23['cdl'] not in cdl_1_2_5 and data_23['numero_studenti_2'] != 0 or data_23['numero_studenti_3'] != 0): 
        pdf.add_page()
        pdf.bg('p13', 'png_cdls')
        report_23_page_2(pdf, data_23)

    # P14
    pdf.add_page()
    pdf.bg('p14', 'png_cdls')

    # P15
    pdf.add_page()
    pdf.bg('p15', 'png_cdls')

    # P16
    pdf.add_page()
    pdf.bg('p16', 'png_cdls')

    # P17
    pdf.add_page()
    pdf.bg('p17', 'png_cdls')

    name = re.sub(r'[^\w\s]', '', cdl).lower().replace(" ", "_")[:-1]

    pdf.output(f"tmp{os.sep}cdls{os.sep}report_{name}.pdf",'F')

def create_pdf_dips(file_1, file_23, dip):
    # parte dei lab

    data_1 = get_exported_data_dips_1(file_1, dip)
    data_23 = get_exported_data_dips_23(file_23, dip)

    pdf = PDF()
    pdf.add_fonts()

    # P1
    pdf.add_page()
    pdf.bg('p1', 'png_dips')
    pdf.set_font('OSb', 'B', 38)
    pdf.set_text_color(54, 77, 98)
    pdf.set_xy(8,199)     
    pdf.multi_cell(0,16, align='L',txt=
                   f"REPORT CONCLUSIVO \n{clean(dip)}")

    # P2
    pdf.add_page()
    pdf.bg('p2', 'png_dips')
    report_2_dip(pdf, data_1, data_23, dip)

    # P3
    pdf.add_page()
    pdf.bg('p3', 'png_dips')

    # P4
    pdf.add_page()
    pdf.bg('p4', 'png_dips')

    # P5
    pdf.add_page()
    pdf.bg('p5', 'png_dips')

    # P6
    pdf.add_page()
    pdf.bg('p6', 'png_dips')

    # P7
    pdf.add_page()
    pdf.bg('p7', 'png_dips')
    report_1_dip(pdf, data_1)

    # P8
    pdf.add_page()
    pdf.bg('p8', 'png_dips')

    # P9
    pdf.add_page()
    pdf.bg('p9', 'png_dips')
    report_3_dip(pdf, data_1)

    # P10
    pdf.add_page()
    pdf.bg('p10', 'png_dips')

    # P11
    pdf.add_page()
    pdf.bg('p11', 'png_dips')
    report_4_dip(pdf, data_1)

    # P12
    pdf.add_page()
    pdf.bg('p12', 'png_dips')

    # P13
    pdf.add_page()
    pdf.bg('p13', 'png_dips')

    # P14
    pdf.add_page()
    pdf.bg('p14', 'png_dips')
    report_23_page_1_dip(pdf, data_23)


    # P15
    pdf.add_page()
    pdf.bg('p15', 'png_dips')
    report_23_page_2_dip(pdf, data_23)
    
    # P16
    pdf.add_page()
    pdf.bg('p16', 'png_dips')
    

    # P17
    pdf.add_page()
    pdf.bg('p17', 'png_dips')

    # P18
    pdf.add_page()
    pdf.bg('p18', 'png_dips')
    report_5_dip(pdf, data_23)

    # P19
    pdf.add_page()
    pdf.bg('p19', 'png_dips')

    # P20
    pdf.add_page()
    pdf.bg('p20', 'png_dips')

    name = dip.lower()
    pdf.output(f"tmp{os.sep}dips{os.sep}report_{name}.pdf",'F')

def create_pdf_pot(file_pot, cdl_pot, uni_pot):
    data_pot = get_exported_data_pot(file_pot, cdl_pot, uni_pot)

    pdf = PDF()
    pdf.add_fonts()

    # P1
    pdf.add_page()
    pdf.bg('p1', 'png_pot_cdls')
    pdf.set_xy(21, 230)     
    pdf.set_font('OSb', 'B', 28)
    pdf.set_text_color(15, 118, 78)

    pdf.multi_cell(170, 12, align='C', txt=f"REPORT CONCLUSIVO DEL CORSO\nDI STUDIO in {clean(cdl_pot)}")

    # P2
    pdf.add_page()
    pdf.bg('p2', 'png_pot_cdls')

    # P3
    pdf.add_page()
    pdf.bg('p3', 'png_pot_cdls')

    # P4
    pdf.add_page()
    pdf.bg('p4', 'png_pot_cdls')

    # P5
    pdf.add_page()
    pdf.bg('p5', 'png_pot_cdls')

    # P6
    pdf.add_page()
    pdf.bg('p6', 'png_pot_cdls')

    # P7
    pdf.add_page()
    pdf.bg('p7', 'png_pot_cdls')
    report_pot(pdf, data_pot)

    # P8
    pdf.add_page()
    pdf.bg('p8', 'png_pot_cdls')
    if data_pot['numero_studenti_pot'] != 0:
        report_table(pdf, data_pot)

    # P9
    pdf.add_page()
    pdf.bg('p9', 'png_pot_cdls')

    name = re.sub(r'[^\w\s]', '', clean(cdl_pot)).lower().replace(" ", "_")
    uni = re.sub(r'[^\w\s]', '', clean(uni_pot)).lower().replace(" ", "_")

    pdf.output(f"tmp{os.sep}pot_cdl{os.sep}report_{name}_{uni}.pdf",'F')

def create_pdf_pot_uni(file_pot, uni_pot):
    data_pot = get_exported_data_pot_uni(file_pot, uni_pot)

    pdf = PDF()
    pdf.add_fonts()

    # P1
    pdf.add_page()
    pdf.bg('p1', 'png_pot_uni')
    pdf.set_xy(21, 230)     
    pdf.set_font('OSb', 'B', 28)
    pdf.set_text_color(15, 118, 78)

    pdf.multi_cell(170, 12, align='C', txt=f"REPORT CONCLUSIVO DELL'{clean(uni_pot)}")

    # P2
    pdf.add_page()
    pdf.bg('p2', 'png_pot_uni')

    # P3
    pdf.add_page()
    pdf.bg('p3', 'png_pot_uni')

    # P4
    pdf.add_page()
    pdf.bg('p4', 'png_pot_uni')

    # P5
    pdf.add_page()
    pdf.bg('p5', 'png_pot_uni')

    # P6
    pdf.add_page()
    pdf.bg('p6', 'png_pot_uni')
    report_pot_uni(pdf, data_pot)

    # P7
    pdf.add_page()
    pdf.bg('p7', 'png_pot_uni')
    
    name = re.sub(r'[^\w\s]', '', clean(uni_pot)).lower().replace(" ", "_")

    pdf.output(f"tmp{os.sep}pot_uni{os.sep}report_{name}.pdf",'F')

############################################################################################################
# Layout

process_encoded = base64.b64encode(open('images' + os.sep + 'buttons' + os.sep + 'process.png', 'rb').read())
select_encoded = base64.b64encode(open('images' + os.sep + 'buttons' + os.sep + 'select.png', 'rb').read())

logo = sg.Image(data=get_img_data('images' + os.sep + 'logo.png', first=True, maxsize=(180,180)))

# Initial window
def get_layout_initial_window():
    return [
        [sg.Column([[logo]], justification='center')],
        [sg.HorizontalSeparator(color='white', pad=(0, 10))],
        [sg.Text('Select the type of report you want to create:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Combo(['Corsi di laurea', 'Dipartimenti', 'POT (CdL)', 'POT (Uni)'],
                  font=('GothamBold', 10, 'bold'), text_color='black', expand_x=True, enable_events=True, readonly=True,
                  key='REPORT_TYPE')],
    ]

# CdL window
def get_input_cdls_window():
    return [
        [sg.Text('Select the CdL:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Combo(list(cdls_dict.keys()), font=('GothamBold', 10, 'bold'), text_color='black', expand_x=True, enable_events=True,
                  readonly=True, key='CDL')],

        [sg.Text('Select the exported file for the 1st year:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Input(key='FILE_1', change_submits=True), sg.Button('', image_data=select_encoded, key='IN_FILE_1')],

        [sg.Text('Select the exported file for the 2nd and 3rd year:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Input(key='FILE_23', change_submits=True), sg.Button('', image_data=select_encoded, key='IN_FILE_23')],

        [sg.Column([[sg.Button('', image_data=process_encoded, key='CREATE_CDLS')]], justification='center')]
    ]

# Dip window
def get_input_dips_window():
    return [
        [sg.Text('Select the Dipartimento:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Combo(list(dips_dict.keys()), font=('GothamBold', 10, 'bold'), text_color='black', expand_x=True, enable_events=True,
                  readonly=True, key='DIP')],

        [sg.Text('Select the exported file for the 1st year:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Input(key='FILE_1', change_submits=True),
         sg.Button('', image_data=select_encoded, key='IN_FILE_1')],

        [sg.Text('Select the exported file for the 2nd and 3rd year:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Input(key='FILE_23', change_submits=True),
         sg.Button('', image_data=select_encoded, key='IN_FILE_23')],

        [sg.Column([[sg.Button('', image_data=process_encoded, key='CREATE_DIPS')]], justification='center')]
    ]

# POT Cdl window
def get_input_pot_cdl_window():
    return [
        [sg.Text('Select the CdL:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Combo(list(pot_cdl_dict.keys()), font=('GothamBold', 10, 'bold'), text_color='black', expand_x=True, enable_events=True,
                  readonly=True, key='CDL_POT')],

        [sg.Text('Select the University:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Combo(list(pot_uni_dict.keys()), font=('GothamBold', 10, 'bold'), text_color='black', expand_x=True, enable_events=True,
                  readonly=True, key='UNI_POT')],

        [sg.Text('Select the exported file:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Input(key='FILE_POT', change_submits=True),
         sg.Button('', image_data=select_encoded, key='IN_FILE_POT')],

        [sg.Column([[sg.Button('', image_data=process_encoded, key='CREATE_POT_CDL')]], justification='center')]
    ]

# POT Uni window
def get_input_pot_uni_window():
    return [
        [sg.Text('Select the University:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Combo(list(pot_uni_dict.keys()), font=('GothamBold', 10, 'bold'), text_color='black', expand_x=True, enable_events=True,
                  readonly=True, key='UNI_POT')],

        [sg.Text('Select the exported file:', text_color='white', font=('GothamBold', 10, 'bold')),
         sg.Input(key='FILE_POT', change_submits=True),
         sg.Button('', image_data=select_encoded, key='IN_FILE_POT')],

        [sg.Column([[sg.Button('', image_data=process_encoded, key='CREATE_POT_UNI')]], justification='center')]
    ]

# Layout Report
def get_layout_window():
    return {
        'Corsi di laurea': [
            [sg.Text("Corsi di laurea",)], [sg.Button('Back')], get_input_cdls_window()
        ],
        'Dipartimenti': [
            [sg.Text("Dipartimenti",)], [sg.Button('Back')], get_input_dips_window()
        ],
        'POT (CdL)': [
            [sg.Text("POT (CdL)",)], [sg.Button('Back')], get_input_pot_cdl_window()
        ],
        'POT (Uni)': [
            [sg.Text("POT (Uni)",)], [sg.Button('Back')], get_input_pot_uni_window()
        ]
    }
